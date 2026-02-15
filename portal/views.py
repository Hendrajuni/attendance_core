from django.shortcuts import render, get_object_or_404, redirect
from datetime import datetime, time, date, timedelta
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.db import transaction
from attendance.utils import get_employee_schedule
from attendance.models import WorkLocation, MonthlyReport, Employee, AttendanceLog, EmployeeLeave
from django.http import HttpResponse
from django.urls import reverse
import calendar
import holidays
from django.utils import timezone
from django.utils.timezone import localtime

# Constants
SCHEDULE_IN = time(8, 0)
SCHEDULE_OUT = time(17, 0)
DEFAULT_SCHEDULE_IN = SCHEDULE_IN
DEFAULT_SCHEDULE_OUT = SCHEDULE_OUT

@login_required
def dashboard(request):
    """
    Main Dashboard View with Role-Based Context
    """
    from attendance.models import WorkLocation, MonthlyReport, Employee, ReportHistory, EmployeeProfile
    from django.db.models import Count, Q, F
    
    context = {
        'page_title': 'Dashboard',
        'current_date': timezone.now().date(),
        'current_month_name': timezone.now().strftime('%B %Y'),
    }

    # Detect Role
    user_role = 'USER'
    user_location = None
    
    if request.user.is_superuser:
        user_role = 'SUPERUSER'
    elif hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        user_role = profile.role
        user_location = profile.assigned_location
        # Fallback to home base if no assigned location
        # Note: EmployeeProfile currently does not link to Employee model
        if not user_location:
             pass # user_location stays None

        # Accounting (Payroll) - New Logic
        if user_role == 'ACCOUNTING':
            # 1. Verified & Pending Counts
            verified_count = MonthlyReport.objects.filter(status='VERIFIED').count()
            pending_count = MonthlyReport.objects.exclude(status='VERIFIED').count()
            
            # 2. Payroll Queue (All Reports, Verified First)
            # Use Case/When to order Verified first
            from django.db.models import Case, When, Value, IntegerField
            payroll_queue = MonthlyReport.objects.annotate(
                status_order=Case(
                    When(status='VERIFIED', then=Value(1)),
                    default=Value(2),
                    output_field=IntegerField(),
                )
            ).exclude(location__parent__isnull=True).order_by('status_order', '-period_year', '-period_month')
            
            context['verified_count'] = verified_count
            context['pending_count'] = pending_count
            context['payroll_queue'] = payroll_queue

        # Kerani
        elif user_role == 'KERANI':
            if user_location:
                # Monthly Report Status for Kerani's location
                try:
                    current_report = MonthlyReport.objects.get(
                        location=user_location,
                        period_year=context['current_date'].year,
                        period_month=context['current_date'].month
                    )
                    context['current_report_status'] = current_report.get_status_display()
                    context['current_report_id'] = current_report.id
                    context['current_report_obj'] = current_report # Pass object for condition checks
                except MonthlyReport.DoesNotExist:
                    context['current_report_status'] = "Belum Dibuat"
                    context['current_report_id'] = None
                    context['current_report_obj'] = None

                # My Activity (Recent Logs for Kerani's location)
                my_activity = ReportHistory.objects.filter(
                    report__location=user_location
                ).select_related('report', 'actor').order_by('-timestamp')[:5]
                context['my_activity'] = my_activity
                
                # Quick Stats for Kerani's location
                my_stats = WorkLocation.objects.filter(id=user_location.id).annotate(
                    total_employees=Count('home_employees'),
                    active_employees=Count('home_employees', filter=Q(home_employees__is_active=True))
                ).first()
                context['my_stats'] = my_stats
            else:
                context['current_report_status'] = "Lokasi Belum Ditentukan"
                context['my_activity'] = None
                context['my_stats'] = None
    
    # Fallback for Superusers if not logged in via Employee Profile (or explicitly Superuser check)
    elif request.user.is_superuser:
        from django.contrib.auth.models import User
        from attendance.models import FingerprintDevice
        
        # 1. System Stats
        total_users = User.objects.count()
        total_devices = FingerprintDevice.objects.count()
        total_reports = MonthlyReport.objects.count()
        
        # 2. Global Activity Log
        recent_activity = ReportHistory.objects.select_related(
            'report', 'report__location', 'actor'
        ).order_by('-timestamp')[:10]
        
        # 3. Device Status
        device_status = FingerprintDevice.objects.select_related('location').order_by('location__code', 'name')
        
        context['total_users'] = total_users
        context['total_devices'] = total_devices
        context['total_reports'] = total_reports
        context['recent_activity'] = recent_activity
        context['device_status'] = device_status

    context['user_role'] = user_role
    context['user_location'] = user_location

    # =========================================================
    # SICK & LEAVE WIDGET DATA
    # =========================================================
    from attendance.models import AttendanceLog, EmployeeLeave
    
    # Base Filters
    today = timezone.now().date()
    log_filter = Q(timestamp__date=today, status__in=['IZIN', 'SAKIT'])
    leave_filter = Q(start_date__lte=today, end_date__gte=today, leave_type__in=['CUTI', 'CUTI_KHUSUS'])
    
    # RBAC Location Filter (For Kerani, Supervisor, Dept Admin)
    # HRD & Superuser see ALL
    if user_role in ['KERANI', 'SUPERVISOR', 'DEPT_ADMIN'] and user_location:
         allowed_locations = user_location.get_descendants(include_self=True)
         log_filter &= Q(employee__home_base__in=allowed_locations)
         leave_filter &= Q(employee__home_base__in=allowed_locations)
    
    # Queries
    sick_logs = AttendanceLog.objects.filter(log_filter).select_related('employee', 'employee__department', 'employee__home_base')
    active_leaves = EmployeeLeave.objects.filter(leave_filter).select_related('employee', 'employee__department')
    
    # Combine Data
    sick_leave_list = []
    
    # Helper to avoid duplicates if needed (using set of IDs)
    processed_ids = set()

    for log in sick_logs:
        if log.employee_id not in processed_ids:
            sick_leave_list.append({
                'employee': log.employee,
                'status_label': log.get_status_display(),
                'status_code': log.status, # IZIN, SAKIT
                'notes': log.notes,
                'location': log.employee.home_base,
                'start_date': log.timestamp.date(),
                'end_date': log.timestamp.date(),
                'is_leave': False
            })
            processed_ids.add(log.employee_id)
        
    for leave in active_leaves:
        if leave.employee_id not in processed_ids:
            sick_leave_list.append({
                'employee': leave.employee,
                'status_label': leave.get_leave_type_display(),
                'status_code': leave.leave_type, # CUTI, CUTI_KHUSUS
                'notes': leave.notes,
                'location': leave.employee.home_base,
                'start_date': leave.start_date,
                'end_date': leave.end_date,
                'is_leave': True
            })
            processed_ids.add(leave.employee_id)
            
    context['sick_leave_list'] = sick_leave_list
    if user_role in ['HRD', 'SUPERUSER', 'ADMIN', 'ACCOUNTING', 'SUPERVISOR'] or request.user.is_superuser:
        
        # 1. Approval Queue (Submitted Reports)
        approval_qs = MonthlyReport.objects.filter(status='SUBMITTED')
        
        # 2. Unlock Request Queue
        unlock_qs = MonthlyReport.objects.filter(status='REQ_UNLOCK')
        
        # 3. Payment Queue (Verified Reports) - For Supervisor/Accounting
        payment_qs = MonthlyReport.objects.filter(status='VERIFIED')

        # FILTER BY LOCATION FOR SUPERVISOR
        if user_role == 'SUPERVISOR' and user_location:
            allowed_locations = user_location.get_descendants(include_self=True)
            approval_qs = approval_qs.filter(location__in=allowed_locations)
            unlock_qs = unlock_qs.filter(location__in=allowed_locations)
            payment_qs = payment_qs.filter(location__in=allowed_locations)
        elif user_role == 'SUPERVISOR' and not user_location:
            # Supervisor without location sees nothing
            approval_qs = approval_qs.none()
            unlock_qs = unlock_qs.none()
            payment_qs = payment_qs.none()

        context['approval_queue'] = approval_qs.select_related('location').order_by('submitted_at')
        context['unlock_queue'] = unlock_qs.select_related('location').order_by('-updated_at')
        context['payment_queue'] = payment_qs.select_related('location').order_by('-verified_at')

        # 3. Employee Stats per Location + Trend (7 Hari)
        # Filter only leaf nodes or locations with employees
        loc_stats_qs = WorkLocation.objects.filter(rght=F('lft') + 1)
        
        if user_role == 'SUPERVISOR' and user_location:
            loc_stats_qs = loc_stats_qs.filter(id__in=user_location.get_descendants(include_self=True))
            
        location_stats = loc_stats_qs.annotate(
            total_employees=Count('home_employees'),
            active_employees=Count('home_employees', filter=Q(home_employees__is_active=True)),
            inactive_employees=Count('home_employees', filter=Q(home_employees__is_active=False)),
            fingerprint_employees=Count('home_employees', filter=Q(home_employees__attendance_method='FINGERPRINT', home_employees__is_active=True)),
            wa_employees=Count('home_employees', filter=Q(home_employees__attendance_method='WHATSAPP', home_employees__is_active=True))
        ).order_by('name')

        # Calculate 7-Day Trend for each location
        today = timezone.localdate()
        start_date = today - timedelta(days=6)
        date_list = [(start_date + timedelta(days=x)) for x in range(7)]
        
        # Get logs for these locations in date range
        # Group by Location and Date
        trend_logs = AttendanceLog.objects.filter(
            timestamp__date__gte=start_date,
            employee__home_base__in=location_stats
        ).values('employee__home_base', 'timestamp__date').annotate(
            present_rs=Count('employee', distinct=True)
        )
        
        # Transform QuerySet into Dict for easy lookup: {loc_id: {date: count}}
        trend_map = {}
        for log in trend_logs:
            loc_id = log['employee__home_base']
            log_date = log['timestamp__date']
            count = log['present_rs']
            
            if loc_id not in trend_map:
                trend_map[loc_id] = {}
            trend_map[loc_id][log_date] = count
            
        # Attach trend_data to location objects
        for loc in location_stats:
            loc_trend = []
            loc_max = 1 # Avoid division by zero
            
            # Use active_employees as baseline/max, or max of the week? 
            # Let's use active_employees as the "100%" mark, but clamp at 100%.
            # Or use dynamic max based on the data.
            baseline = loc.active_employees if loc.active_employees > 0 else 1

            for day in date_list:
                val = trend_map.get(loc.id, {}).get(day, 0)
                # Calculate percentage relative to active employees
                pct = min(100, int((val / baseline) * 100))
                loc_trend.append({
                    'date': day,
                    'count': val,
                    'pct': pct,
                    'day_name': day.strftime('%a') # Mon, Tue...
                })
            
            loc.trend_data = loc_trend
        
        context['location_stats'] = location_stats
        
        # 4. Global Activity Log
        activity_qs = ReportHistory.objects.select_related('report', 'report__location', 'actor')
        
        if user_role == 'SUPERVISOR' and user_location:
             activity_qs = activity_qs.filter(report__location__in=user_location.get_descendants(include_self=True))
             
        context['recent_activity'] = activity_qs.order_by('-timestamp')[:8]

    # =========================================================
    # Context for KERANI (Location Admin)
    # =========================================================
    # Process for Kerani OR if user has a location assigned (even if HRD/Admin wants to see local data)
    if user_location:
        
        # 1. Current Month Report Status
        current_month = timezone.now().month
        current_year = timezone.now().year
        
        current_report = MonthlyReport.objects.filter(
            location=user_location,
            period_month=current_month,
            period_year=current_year
        ).first()
        
        context['current_report'] = current_report
        
        # 2. My Location Activity Log
        my_activity = ReportHistory.objects.filter(
            report__location=user_location
        ).select_related('actor').order_by('-timestamp')[:8]
        
        context['my_activity'] = my_activity
        
        # 3. Quick Stats
        # Employee count in my location
        my_stats = {
            'total': Employee.objects.filter(home_base=user_location).count(),
            'active': Employee.objects.filter(home_base=user_location, is_active=True).count(),
            'inactive': Employee.objects.filter(home_base=user_location, is_active=False).count(),
        }
        context['my_stats'] = my_stats

    return render(request, 'portal/dashboard.html', context)

@login_required
def attendance_reports_dashboard(request):
    """
    Dashboard Laporan Absensi (Digital Archive).
    Menampilkan status laporan per bulan dalam setahun.
    """
    from attendance.models import WorkLocation, MonthlyReport
    from django.utils import timezone
    from django.db.models import F
    
    # 1. Determine Location (RBAC) - Copied from recap_matrix_view logic
    selected_location = None
    can_select_location = False
    all_locations = WorkLocation.objects.none()

    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.assigned_location:
            selected_location = profile.assigned_location
        
        # Admin / HRD / Accounting can switch locations
        if profile.role in ['ADMIN', 'HRD', 'ACCOUNTING']:
            can_select_location = True
            # Hanya ambil lokasi Leaf Node (PKS, Kebun) yang punya data absensi
            all_locations = WorkLocation.objects.filter(rght=F('lft') + 1).order_by('tree_id', 'lft')
            
            # Allow filter override
            location_id = request.GET.get('location_id')
            if location_id:
                try:
                    selected_location = WorkLocation.objects.get(id=location_id)
                except WorkLocation.DoesNotExist:
                    pass
            
            # Default to first location if none selected (or if assigned_location is not in list)
            if not selected_location and all_locations.exists():
                selected_location = all_locations.first()
    else:
        # Fallback for superuser
        if request.user.is_superuser or request.user.is_staff:
            can_select_location = True
            all_locations = WorkLocation.objects.filter(rght=F('lft') + 1).order_by('tree_id', 'lft')
            location_id = request.GET.get('location_id')
            if location_id:
                try:
                    selected_location = WorkLocation.objects.get(id=location_id)
                except WorkLocation.DoesNotExist:
                    pass
            if not selected_location and all_locations.exists():
                selected_location = all_locations.first()

    # 2. Determine Year
    current_year = timezone.now().year
    try:
        year = int(request.GET.get('year', current_year))
    except (ValueError, TypeError):
        year = current_year
    
    # 3. Build Month Cards Data
    months_data = []
    
    # Pre-fetch reports for query optimization
    reports_map = {}
    wa_months_with_data = set()

    if selected_location:
        reports = MonthlyReport.objects.filter(
            location=selected_location,
            period_year=year
        )
        for r in reports:
            reports_map[r.period_month] = r
            
        # Pre-fetch months that have WA data
        from attendance.models import AttendanceLog
        from django.db.models import Q
        
        # Check logs directly for WA/Telegram indicators, regardless of employee's current setting.
        # Indicators: source_type='TELEGRAM' or verification_method='WA'
        wa_logs = AttendanceLog.objects.filter(
            timestamp__year=year,
            employee__home_base=selected_location
        ).filter(
            Q(source_type='TELEGRAM') | Q(verification_method='WA') | Q(employee__attendance_method__in=['WHATSAPP', 'WA'])
        ).values_list('timestamp__month', flat=True).distinct()
        
        wa_months_with_data = set(wa_logs)
            
    month_names = [
        '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
    ]

    for m in range(1, 13):
        report = reports_map.get(m)
        has_wa_data = m in wa_months_with_data
        
        status_class = 'secondary' # Default empty/gray
        if report:
            if report.status == 'DRAFT':
                status_class = 'warning' # Kuning
            elif report.status == 'SUBMITTED':
                status_class = 'info'    # Biru
            elif report.status == 'VERIFIED':
                status_class = 'success' # Hijau
            elif report.status == 'REQ_REVISI':
                status_class = 'danger'  # Merah
            elif report.status == 'REJECTED':
                status_class = 'danger'
        
        months_data.append({
            'month_num': m,
            'month_name': month_names[m],
            'report': report,
            'status_class': status_class,
            'has_wa_data': has_wa_data
        })

    context = {
        'page_title': 'Dashboard Laporan Absensi',
        'selected_location': selected_location,
        'all_locations': all_locations,
        'can_select_location': can_select_location,
        'year': year,
        'months_data': months_data,
        'years_range': range(current_year - 2, current_year + 2),
    }
    
    return render(request, 'portal/attendance_reports.html', context)

@login_required
def tree_explorer(request):
    """
    Pusat Data Wilayah View (Hierarchy)
    """
    from attendance.models import WorkLocation
    
    all_locations = WorkLocation.objects.all()
    locations = all_locations
    
    # RBAC: Kerani only sees their location and descendants
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role in ['KERANI', 'SUPERVISOR', 'DEPT_ADMIN'] and profile.assigned_location:
            # get_descendants(include_self=True) handles the tree traversal efficiently
            locations = profile.assigned_location.get_descendants(include_self=True)
            
    # Annotate with device and spreadsheet presence
    from django.db.models import Count
    locations = locations.annotate(
        device_count=Count('machine_configs', distinct=True),
        wa_count=Count('spreadsheets', distinct=True)
    )

    context = {
        'page_title': 'Pusat Data Wilayah',
        'locations': locations,
        'all_locations': all_locations,  # For mutation modal dropdown
    }
    return render(request, 'portal/tree_explorer.html', context)


@login_required
def location_detail_view(request, location_id):
    """
    HTMX: Load detail konten untuk panel kanan Tree Explorer
    Supports pagination with page_size for employees and logs
    """
    from attendance.models import WorkLocation, Employee, AttendanceLog
    from django.utils import timezone
    from django.core.paginator import Paginator
    
    location = get_object_or_404(WorkLocation, id=location_id)
    
    # Pagination parameters
    emp_page_size = int(request.GET.get('emp_page_size', 10))  # Default 10
    emp_page = int(request.GET.get('emp_page', 1))
    log_page_size = int(request.GET.get('log_page_size', 10))  # Default 10
    log_page = int(request.GET.get('log_page', 1))
    
    # Valid page sizes
    PAGE_SIZE_OPTIONS = [10, 20, 50, 100]
    if emp_page_size not in PAGE_SIZE_OPTIONS:
        emp_page_size = 10
    if log_page_size not in PAGE_SIZE_OPTIONS:
        log_page_size = 10
    
    # Ambil lokasi ini DAN seluruh turunannya (jika ada)
    # Ini penting untuk melihat total karyawan di wilayah tersebut (termasuk sub-wilayah)
    # Gunakan get_descendants(include_self=True) dari MPTT
    sub_locations = location.get_descendants(include_self=True)
    
    # Filter Karyawan: Home Base is within these locations
    employees_qs = Employee.objects.filter(
        home_base__in=sub_locations,
        is_verified=True
    ).select_related('department', 'home_base').order_by('full_name')
    
    total_employees = employees_qs.count()
    active_employees = employees_qs.filter(is_active=True).count()
    
    # Paginate employees
    emp_paginator = Paginator(employees_qs, emp_page_size)
    employees_page = emp_paginator.get_page(emp_page)
    
    # Statistik Kehadiran Hari Ini (Sederhana)
    today = timezone.now().date()
    # Log hari ini untuk karyawan di wilayah ini
    logs_today = AttendanceLog.objects.filter(
        employee__in=employees_qs,
        timestamp__date=today
    ).values('employee').distinct().count()
    
    # Recent Logs for Tab with pagination
    recent_logs_qs = AttendanceLog.objects.filter(
        employee__in=employees_qs
    ).select_related('employee', 'captured_at').order_by('-timestamp')
    
    log_paginator = Paginator(recent_logs_qs, log_page_size)
    logs_page = log_paginator.get_page(log_page)
    
    context = {
        'location': location,
        'employees': employees_page,
        'employees_page': employees_page,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'present_today': logs_today,
        'recent_logs': logs_page,
        'logs_page': logs_page,
        # Pagination state
        'emp_page_size': emp_page_size,
        'log_page_size': log_page_size,
        'page_size_options': PAGE_SIZE_OPTIONS,
    }
    
    # If HTMX request, render partial (for tree explorer panel)
    if request.headers.get('HX-Request'):
        return render(request, 'portal/partials/_location_detail.html', context)
    
    # If standard request, render full page wrapper
    return render(request, 'portal/location_detail.html', context)

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
from attendance.models import FingerprintDevice, SpreadsheetSource, AttendanceLog, Employee, WorkLocation, EmployeeProfile, EmployeeLeave

@login_required
def sync_logs_view(request):
    """
    Pusat Sinkronisasi View with RBAC
    """
    devices = FingerprintDevice.objects.filter(is_active=True).select_related('location')
    sources = SpreadsheetSource.objects.all().select_related('location')
    
    # RBAC Filter: Kerani only sees their location
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location:
            devices = devices.filter(location=profile.assigned_location)
            sources = sources.filter(location=profile.assigned_location)
            
    context = {
        'page_title': 'Tarik Log Absensi',
        'devices': devices,
        'sources': sources,
    }
    return render(request, 'portal/sync_center.html', context)


@login_required
def sync_employees_view(request):
    """
    Tarik Data Karyawan Baru View
    """
    devices = FingerprintDevice.objects.filter(is_active=True).select_related('location')
    sources = SpreadsheetSource.objects.all().select_related('location')
    
    # RBAC Filter
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location:
            devices = devices.filter(location=profile.assigned_location)
            sources = sources.filter(location=profile.assigned_location)
            
    context = {
        'page_title': 'Tarik Karyawan Baru',
        'devices': devices,
        'sources': sources,
        'is_employee_sync': True, # Flag for template matching
    }
    return render(request, 'portal/sync_employees.html', context)


@login_required
@require_POST
def sync_employee_machine_htmx(request, device_id):
    device = get_object_or_404(FingerprintDevice, id=device_id)
    
    # Simple Access Check
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location != device.location:
             return HttpResponse('<span class="badge bg-danger">Akses Ditolak</span>')

    from attendance.utils import fetch_users_from_machine
    
    count, samples, error = fetch_users_from_machine(device)
    
    if error:
        return HttpResponse(f'''
            <span class="badge bg-danger">Error</span>
            <tr hx-swap-oob="afterbegin:#sync-log-body">
                <td>{timezone.now().strftime("%H:%M:%S")}</td>
                <td>{device.name}</td>
                <td><span class="badge bg-danger">Gagal</span></td>
                <td class="text-danger">{error}</td>
            </tr>
        ''')
    
    # ...
    
    current_time = timezone.now().strftime("%H:%M:%S")
    
    if count > 0:
        badge_html = f'<span class="badge bg-success">User: {count}</span> <span class="badge bg-light text-dark border">Cek: {current_time}</span>'
        msg_text = f"Berhasil menarik {count} user baru"
    else:
        badge_html = f'<span class="badge bg-secondary">User: 0</span> <span class="badge bg-light text-dark border">Cek: {current_time}</span>'
        msg_text = "Tidak ada user baru (Up-to-date)"
        
    # Build content for SyncLogManager (JS)
    # We will pass a single string containing all new TRs to the JS function
    new_rows_html = ""
    
    if samples:
        for sample in samples:
            new_rows_html += f'''
            <tr>
                <td>{current_time}</td>
                <td>{device.name}</td>
                <td><span class="badge bg-info">Detail</span></td>
                <td class="text-info">{sample}</td>
            </tr>
            '''
    
    new_rows_html += f'''
        <tr>
            <td>{current_time}</td>
            <td>{device.name}</td>
            <td><span class="badge bg-{'success' if count > 0 else 'secondary'}">Sukses</span></td>
            <td>{msg_text}</td>
        </tr>
    '''
    
    # Escape inner quotes for JS string
    new_rows_html = new_rows_html.replace('\n', '').replace('"', "'")
    
    # Return badge update (standard HTMX) + Script to update logs via Manager
    return HttpResponse(f'''
        {badge_html}
        <script>
            if (typeof SyncLogManager !== 'undefined') {{
                SyncLogManager.addRows("{new_rows_html}");
            }}
        </script>
    ''')


@login_required
@require_POST
def sync_employee_wa_htmx(request, source_id):
    source = get_object_or_404(SpreadsheetSource, id=source_id)
    
    # Simple Access Check
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location != source.location:
             return HttpResponse('<span class="badge bg-danger">Akses Ditolak</span>')

    from attendance.utils import fetch_users_from_wa_source
    
    count, samples, error = fetch_users_from_wa_source(source)
    
    if error:
         return HttpResponse(f'''
            <span id="status-source-{source.id}" class="badge bg-danger" hx-swap-oob="true">Error</span>
             <tr hx-swap-oob="afterbegin:#sync-log-body">
                <td>{timezone.now().strftime("%H:%M:%S")}</td>
                <td>{source.name}</td>
                <td><span class="badge bg-danger">Gagal</span></td>
                <td class="text-danger">{error}</td>
            </tr>
        ''')

    # Main Status Update (OOB) to avoid messy main-swap context
    status_html = f'<span id="status-source-{source.id}" class="badge bg-success" hx-swap-oob="true">OK ({count})</span>'
    
    # Build Detail Rows for Samples
    extra_rows = ""
    current_time = timezone.now().strftime("%H:%M:%S")
    if samples:
        for sample in samples:
            extra_rows += f'''
            <tr hx-swap-oob="afterbegin:#sync-log-body">
                <td>{current_time}</td>
                <td>{source.name}</td>
                <td><span class="badge bg-info">Detail</span></td>
                <td class="text-info">{sample}</td>
            </tr>
            '''
            
    # Main Log Row
    log_row = f'''
        <tr hx-swap-oob="afterbegin:#sync-log-body">
            <td>{timezone.now().strftime("%H:%M:%S")}</td>
            <td>{source.name}</td>
            <td><span class="badge bg-success">Sukses</span></td>
            <td>Berhasil menarik {count} karyawan baru</td>
        </tr>
    '''
    
    # Return everything combined.
    # ALTERNATIVE APPROACH: JS Injection
    # Since HTMX/Browser parsing of OOB <tr> tags is flaky, we use a simple script to inject the rows.
    # This bypasses the parsing "tr must be inside table" rule because it's treated as a JS string until insertion.
    
    # Escape single quotes and newlines for JS string safety
    rows_js = (extra_rows + log_row).replace('"', '\\"').replace('\n', '')
    
    response_html = f'''
    {status_html.replace('hx-swap-oob="true"', '')} <!-- Return normal HTML for the target -->
    <script>
        var newRows = "{rows_js}";
        // Remove hx-swap-oob attributes since we are injecting manually
        newRows = newRows.replace(/hx-swap-oob="[^"]+"/g, "");
        
        // Use our new Client-Side Manager
        if (typeof SyncLogManager !== 'undefined') {{
            SyncLogManager.addRows(newRows);
        }} else {{
            // Fallback if script not loaded yet (should rarely happen)
            document.querySelector('#sync-log-body').insertAdjacentHTML('afterbegin', newRows);
        }}
    </script>
    '''
    return HttpResponse(response_html)


@login_required
@require_POST
def sync_machine_htmx(request, device_id):
    device = get_object_or_404(FingerprintDevice, id=device_id)
    print(f"DEBUG: HIT sync_machine_htmx for {device.name} (IP: {device.ip_address})")
    
    # Simple Access Check
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location != device.location:
             return HttpResponse('<span class="badge bg-danger">Akses Ditolak</span>')

    try:
        from zk import ZK
        from attendance.utils import determine_category
        from portal.models import Notification
        
        print("DEBUG: Connecting to ZK...")
        zk = ZK(device.ip_address, port=device.port, timeout=20)
        conn = zk.connect()
        
        if not conn:
             return HttpResponse(f'''
                <span class="badge bg-danger">Gagal Koneksi</span>
                <tr hx-swap-oob="afterbegin:#sync-log-body">
                    <td>{timezone.now().strftime("%H:%M:%S")}</td>
                    <td>{device.name}</td>
                    <td><span class="badge bg-danger">Error</span></td>
                    <td class="text-danger">Gagal terhubung ke mesin (Timeout)</td>
                </tr>
            ''')
            
        try:
            print("DEBUG: Fetching attendance...")
            attendances = conn.get_attendance()
            total_records = len(attendances)
            print(f"DEBUG: Fetched {total_records} records. Optimizing processing...")
            
            created_count = 0
            
            # OPTIMIZATION: Prefetch all active employees to avoid N+1 queries
            # Map device_user_id -> Employee Object
            active_employees = {
                str(emp.device_user_id): emp 
                for emp in Employee.objects.filter(is_active=True)
            }
            print(f"DEBUG: Loaded {len(active_employees)} active employees into memory.")

            # Track detailed logs for UI (Limit 5)
            detailed_logs = []
            
            with transaction.atomic():
                for att in attendances:
                    user_id = str(att.user_id)
                    
                    # Fast Lookup from Memory
                    employee = active_employees.get(user_id)
                    if not employee:
                        continue
                        
                    timestamp = att.timestamp
                    if timezone.is_naive(timestamp):
                        timestamp = timezone.make_aware(timestamp)
                    
                    # Determine category (CPU bound, fast)
                    category, _ = determine_category(employee, timestamp)
                    
                    # We still use get_or_create for safety, but now 'employee' is already fetched
                    _, created = AttendanceLog.objects.get_or_create(
                        employee=employee, timestamp=timestamp,
                        defaults={
                            'status': 'HADIR', 'source_type': 'FINGERPRINT',
                            'verification_method': 'FINGER', 'captured_at': device.location,
                            'log_category': category
                        }
                    )
                    if created: 
                        created_count += 1
                        # Store tuple: (name, time_str, category) for rich display
                        time_str = timestamp.strftime("%H:%M")
                        detailed_logs.append((employee.full_name, time_str, category))
            
            print(f"DEBUG: Created {created_count} new logs from {total_records} raw records.")
            current_time = timezone.now().strftime("%H:%M:%S")
            
            # --- Update Last Activity ---
            device.last_sync = timezone.now()
            device.save(update_fields=['last_sync'])
            
            # OOB Status Update
            status_html = f'<span id="status-device-{device.id}" class="badge bg-success" hx-swap-oob="true">OK ({created_count})</span>'
            
            # --- Create Notification (ALWAYS) ---
            from portal.models import Notification
            from attendance.models import EmployeeProfile 
            from django.contrib.auth.models import User
            
            actor_name = request.user.first_name or request.user.username
            time_str = timezone.localtime(timezone.now()).strftime("%d/%m %H:%M")
            
            title = f"Sync: {device.name} ({device.location.code})"
            if created_count > 0:
                message = f"Ditarik oleh {actor_name} pada {time_str}. Total {created_count} data baru dari {device.location.name}."
            else:
                message = f"Dicek oleh {actor_name} pada {time_str}. Tidak ada data baru dari {device.location.name}."
            
            # Logic: Send to Admin/HRD (Global) + Manager/Kerani (Location Specific)
            # 1. Global (Admin/HRD) + All Superusers
            global_recipients = set(EmployeeProfile.objects.filter(role__in=['ADMIN', 'HRD']).values_list('user', flat=True))
            superusers = set(User.objects.filter(is_superuser=True).values_list('id', flat=True))
            
            # 2. Local (Manager/Kerani at this location)
            local_recipients = set(EmployeeProfile.objects.filter(
                role__in=['HRD', 'KERANI'], 
                assigned_location=device.location
            ).values_list('user', flat=True))
            
            recipient_ids = global_recipients.union(superusers).union(local_recipients)
            
            notifications_to_create = []
            for uid in recipient_ids:
                notifications_to_create.append(Notification(
                    recipient_id=uid,
                    title=title,
                    message=message,
                    related_location=device.location
                ))
            
            Notification.objects.bulk_create(notifications_to_create)

            # Badge logic
            if created_count > 0:
                badge_html = f'<span class="badge bg-success">Data: {created_count}</span> <span class="badge bg-light text-dark border">Cek: {current_time}</span>'
                msg_text = f"Berhasil menarik {created_count} data baru (Total Mesin: {total_records})"
            else:
                badge_html = f'<span class="badge bg-secondary">Data: 0</span> <span class="badge bg-light text-dark border">Cek: {current_time}</span>'
                msg_text = f"Tidak ada data baru. (Total di Mesin: {total_records})"

           # Build HTML for new logs (Script Injection Method)
            extra_rows = ""
            if detailed_logs:
                for name, att_time, category in detailed_logs:
                    # Dynamic badge based on category
                    if category in ('CHECKIN_1', 'SCAN_IN'):
                        badge = '<span class="badge bg-success">Masuk</span>'
                        text_class = 'text-success'
                    elif category in ('CHECKOUT', 'SCAN_OUT'):
                        badge = '<span class="badge bg-primary">Pulang</span>'
                        text_class = 'text-primary'
                    elif category == 'CHECKIN_2':
                        badge = '<span class="badge bg-info">Cek-2</span>'
                        text_class = 'text-info'
                    else:
                        badge = '<span class="badge bg-secondary">Absen</span>'
                        text_class = 'text-light'
                    
                    extra_rows += f'''
                    <tr>
                        <td>{att_time}</td>
                        <td>{device.name}</td>
                        <td>{badge}</td>
                        <td class="{text_class}">{name}</td>
                    </tr>
                    '''
            
            log_row = f'''
                <tr>
                    <td>{current_time}</td>
                    <td>{device.name}</td>
                    <td><span class="badge bg-success">Sukses</span></td>
                    <td>Berhasil menarik {created_count} data baru (Total Mesin: {total_records})</td>
                </tr>
            '''
            
            # Escape strings for JS safely
            import json
            rows_js = json.dumps(extra_rows + log_row)
            
            status_html = f'<span id="status-device-{device.id}" class="badge bg-success">OK ({total_records})</span>'
            
            return HttpResponse(f'''
                {status_html}
                <script>
                    var newRows = {rows_js};
                    if (window.SyncLogManager) {{
                         window.SyncLogManager.addRows(newRows);
                    }} else {{
                         // Fallback
                         document.querySelector('#sync-log-body').insertAdjacentHTML('afterbegin', newRows);
                    }}
                </script>
            ''')
            
        except Exception as e:
            print(f"DEBUG: Error during sync: {e}")
            import traceback
            traceback.print_exc()
            
            import json
            error_msg = str(e)
            error_row_html = f'''
                <tr>
                    <td>{timezone.now().strftime("%H:%M:%S")}</td>
                    <td>{device.name}</td>
                    <td><span class="badge bg-danger">Error</span></td>
                    <td class="text-danger">{error_msg}</td>
                </tr>
            '''
            rows_js = json.dumps(error_row_html)

            return HttpResponse(f'''
                <span class="badge bg-danger">Error</span>
                 <script>
                    var errorRows = {rows_js};
                    if (window.SyncLogManager) {{
                         window.SyncLogManager.addRows(errorRows);
                    }} else {{
                         document.querySelector('#sync-log-body').insertAdjacentHTML('afterbegin', errorRows);
                    }}
                </script>
            ''')
            
        finally:
            conn.disconnect()

    except Exception as e:
        print(f"DEBUG: Top level error: {e}")
        return HttpResponse(f'<span class="badge bg-danger">Error Connect</span>')


@login_required
def ping_machine_htmx(request, device_id):
    device = get_object_or_404(FingerprintDevice, id=device_id)
    
    try:
        from zk import ZK
        # Increase timeout to 10s matching debug script
        zk = ZK(device.ip_address, port=device.port, timeout=10)
        conn = zk.connect()
        if conn:
            conn.disconnect()
            device.last_sync = timezone.now()
            device.save(update_fields=['last_sync'])
            return HttpResponse('<span class="badge bg-success">Online</span>')
        else:
            return HttpResponse('<span class="badge bg-danger">Offline (No Conn)</span>')
    except Exception as e:
        print(f"PING ERROR {device.ip_address}: {str(e)}") # Log error to console
        return HttpResponse(f'<span class="badge bg-danger" title="{str(e)}">Offline</span>')

@login_required
@require_POST
def sync_wa_source_htmx(request, source_id):
    source = get_object_or_404(SpreadsheetSource, id=source_id)
    
     # Simple Access Check
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role == 'KERANI' and profile.assigned_location != source.location:
             return HttpResponse('<span class="badge bg-danger">Akses Ditolak</span>')

    try:
        import pandas as pd
        from attendance.utils import determine_category
        
        csv_url = f"https://docs.google.com/spreadsheets/d/{source.spreadsheet_id}/gviz/tq?tqx=out:csv&sheet={source.sheet_name}"
        df = pd.read_csv(csv_url)
        df.columns = [c.strip().upper() for c in df.columns]
        
        # Track detailed logs for UI
        detailed_logs = []
        created_count = 0
        
        # DEBUG: Print column names
        print(f"DEBUG: Spreadsheet columns = {list(df.columns)}")
        print(f"DEBUG: Total rows in spreadsheet = {len(df)}")
        
        with transaction.atomic():
            for idx, row in df.iterrows():
                tgl_str = str(row.get('TANGGAL', '')).strip()
                jam_str = str(row.get('JAM ABSEN', '')).strip()
                user_id = str(row.get('NOMOR WA', '')).strip()
                
                if not tgl_str or not jam_str: 
                    print(f"DEBUG: Row {idx} skipped - missing TANGGAL or JAM ABSEN")
                    continue
                
                try:
                    full_ts_str = f"{tgl_str} {jam_str}"
                    timestamp = pd.to_datetime(full_ts_str, dayfirst=True).to_pydatetime()
                    if timezone.is_naive(timestamp):
                        timestamp = timezone.make_aware(timestamp)
                    
                    # Normalize phone number lookup (handles .0 suffix from float storage)
                    # Try exact match first
                    employee = Employee.objects.filter(phone_number=user_id, is_active=True).first()
                    if not employee:
                        # Try with .0 suffix (database might have stored as float)
                        employee = Employee.objects.filter(phone_number=f"{user_id}.0", is_active=True).first()
                    if not employee:
                        employee = Employee.objects.filter(telegram_user_id=user_id, is_active=True).first()
                    if not employee:
                        # Try telegram_user_id with .0 suffix
                        employee = Employee.objects.filter(telegram_user_id=f"{user_id}.0", is_active=True).first()
                    if not employee: 
                        print(f"DEBUG: Row {idx} - Employee not found for user_id={user_id}")
                        continue
                    
                    category, _ = determine_category(employee, timestamp)
                    
                    _, created = AttendanceLog.objects.get_or_create(
                         employee=employee, timestamp=timestamp,
                         defaults={
                            'status': 'HADIR', 'source_type': 'TELEGRAM',
                            'verification_method': 'WA', 'captured_at': source.location,
                            'log_category': category
                        }
                    )
                    if created: 
                        created_count += 1
                        # Store tuple: (name, time_str, category) for rich display
                        time_str = timestamp.strftime("%H:%M")
                        detailed_logs.append((employee.full_name, time_str, category))
                    else:
                        print(f"DEBUG: Row {idx} - Log already exists for {employee.full_name} at {timestamp}")
                except Exception as row_err:
                    print(f"DEBUG: Row {idx} - Error: {row_err}")
                    continue
        
        print(f"DEBUG: Created {created_count} new logs")
        
        # --- Update Last Activity ---
        source.last_activity = timezone.now()
        source.save(update_fields=['last_activity'])
        
        # --- Create Notification (ALWAYS) ---
        from portal.models import Notification
        from attendance.models import EmployeeProfile
        from django.contrib.auth.models import User
        
        actor_name = request.user.first_name or request.user.username
        time_str = timezone.localtime(timezone.now()).strftime("%d/%m %H:%M")
        
        title = f"Sync WA: {source.name} ({source.location.code})"
        if created_count > 0:
            message = f"Ditarik oleh {actor_name} pada {time_str}. Total {created_count} data baru dari {source.location.name}."
        else:
            message = f"Dicek oleh {actor_name} pada {time_str}. Tidak ada data baru dari {source.location.name}."
        
        # Logic: Send to Admin/HRD (Global) + Manager/Kerani (Location Specific)
        global_recipients = set(EmployeeProfile.objects.filter(role__in=['ADMIN', 'HRD']).values_list('user', flat=True))
        superusers = set(User.objects.filter(is_superuser=True).values_list('id', flat=True))

        local_recipients = set(EmployeeProfile.objects.filter(
            role__in=['HRD', 'KERANI'],
            assigned_location=source.location
        ).values_list('user', flat=True))
        
        recipient_ids = global_recipients.union(superusers).union(local_recipients)
        
        notifications_to_create = []
        for uid in recipient_ids:
            notifications_to_create.append(Notification(
                recipient_id=uid,
                title=title,
                message=message,
                related_location=source.location
            ))
        
        Notification.objects.bulk_create(notifications_to_create)

        status_badge = f'<span class="badge bg-success">OK ({created_count})</span>'
        
        # Build HTML for new logs (Script Injection Method)
        import json
        extra_rows = ""
        current_time = timezone.now().strftime("%H:%M:%S")
        
        for name, att_time, category in detailed_logs:
            # Dynamic badge based on category
            if category in ('CHECKIN_1', 'SCAN_IN'):
                badge = '<span class="badge bg-success">Masuk</span>'
                text_class = 'text-success'
            elif category in ('CHECKOUT', 'SCAN_OUT'):
                badge = '<span class="badge bg-primary">Pulang</span>'
                text_class = 'text-primary'
            elif category == 'CHECKIN_2':
                badge = '<span class="badge bg-info">Cek-2</span>'
                text_class = 'text-info'
            else:
                badge = '<span class="badge bg-secondary">Absen</span>'
                text_class = 'text-light'
            
            extra_rows += f'''
            <tr>
                <td>{att_time}</td>
                <td>{source.name}</td>
                <td>{badge}</td>
                <td class="{text_class}">{name}</td>
            </tr>
            '''
        
        log_row = f'''
            <tr>
                <td>{current_time}</td>
                <td>{source.name}</td>
                <td><span class="badge bg-success">Sukses</span></td>
                <td>Berhasil menarik {created_count} data baru</td>
            </tr>
        '''
        
        rows_js = json.dumps(extra_rows + log_row)

        return HttpResponse(f'''
            {status_badge}
            <script>
                var newRows = {rows_js};
                if (window.SyncLogManager) {{
                     window.SyncLogManager.addRows(newRows);
                }} else {{
                     // Fallback
                     document.querySelector('#sync-log-body').insertAdjacentHTML('afterbegin', newRows);
                }}
            </script>
        ''')

    except Exception as e:
        import json
        error_msg = str(e)[:100]
        error_row_html = f'''
            <tr>
                <td>{timezone.now().strftime("%H:%M:%S")}</td>
                <td>{source.name}</td>
                <td><span class="badge bg-danger">Exception</span></td>
                <td class="text-danger">{error_msg}</td>
            </tr>
        '''
        rows_js = json.dumps(error_row_html)
        
        return HttpResponse(f'''
            <span class="badge bg-danger">Error</span>
             <script>
                var errorRows = {rows_js};
                if (window.SyncLogManager) {{
                     window.SyncLogManager.addRows(errorRows);
                }} else {{
                     document.querySelector('#sync-log-body').insertAdjacentHTML('afterbegin', errorRows);
                }}
            </script>
        ''')

@login_required
def reports(request):
    """
    Laporan View (Restricted Access)
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return render(request, 'portal/403.html', status=403)
        
    context = {
        'page_title': 'Laporan',
    }
    return render(request, 'portal/reports.html', context)

@login_required
def settings(request):
    """
    Pengaturan View
    """
    context = {
        'page_title': 'Pengaturan',
    }
    return render(request, 'portal/settings.html', context)


@login_required
def global_search(request):
    """
    Global Search View (HTMX)
    Searches Employees by name or NIK
    """
    query = request.GET.get('q', '').strip()
    
    if not query:
        return HttpResponse("")
        
    results = Employee.objects.filter(
        Q(full_name__icontains=query) | Q(nik__icontains=query)
    )[:5] # Limit to 5 results
    
    if not results.exists():
        return HttpResponse('<li class="dropdown-item text-muted small">Tidak ditemukan</li>')
        
    html = ""
    for emp in results:
        # Determine status color
        status_color = "success" if emp.is_active else "secondary"
        
        html += f'''
        <li>
            <a class="dropdown-item d-flex align-items-center gap-2 py-2" href="/admin/attendance/employee/{emp.id}/change/">
                <div class="bg-light rounded-circle d-flex align-items-center justify-content-center" style="width: 32px; height: 32px;">
                    <i class="fas fa-user text-secondary small"></i>
                </div>
                <div>
                    <div class="fw-medium small">{emp.full_name}</div>
                    <div class="text-muted" style="font-size: 0.75rem;">
                        {emp.nik or "No NIK"} <span class="badge bg-{status_color} ms-1" style="font-size: 0.6rem;">{emp.get_status_display() if hasattr(emp, 'get_status_display') else 'Active'}</span>
                    </div>
                </div>
            </a>
        </li>
        '''
        
    
    # Add "View All" link if needed, or just return list
    html += '<li><hr class="dropdown-divider"></li>'
    # Link to Admin list filtered by query is a nice touch
    admin_search_url = f"/admin/attendance/employee/?q={query}"
    html += f'<li><a class="dropdown-item text-center small text-primary fw-bold" href="{admin_search_url}">Lihat Semua Hasil</a></li>'
    
    return HttpResponse(html)

@login_required
@require_POST
def mark_notification_read_htmx(request, notification_id):
    """
    Mark notification as read and return updated UI
    """
    from .models import Notification
    
    notif = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    if not notif.is_read:
        notif.is_read = True
        notif.save()
        
    unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    # Badge OOB Update
    badge_html = ""
    if unread_count > 0:
         badge_html = f'''
         <span id="notif-badge-count" class="position-absolute top-0 start-100 translate-middle p-1 bg-danger border border-light rounded-circle" hx-swap-oob="true">
            <span class="visually-hidden">New alerts</span>
         </span>
         <span id="notif-header-count" class="badge bg-danger rounded-pill" hx-swap-oob="true">{unread_count} Baru</span>
         '''
    else:
        # Remove badge if 0
         badge_html = f'''
         <span id="notif-badge-count" class="d-none" hx-swap-oob="true"></span>
         <span id="notif-header-count" class="d-none" hx-swap-oob="true"></span>
         '''

    # Return the updated item (without bg-light and check button changed)
    # We reconstruct the LI here
    item_html = f'''
    <li id="notif-item-{notif.id}">
        <div class="dropdown-item small py-2 d-flex justify-content-between align-items-start">
            <a href="#" class="text-decoration-none text-dark flex-grow-1">
                <div class="d-flex align-items-start">
                    <i class="fas fa-info-circle text-secondary mt-1 me-2"></i>
                    <div>
                        <div class="fw-medium">{notif.title}</div>
                        <div class="text-muted text-truncate" style="max-width: 250px;">{notif.message}</div>
                        <div class="text-muted" style="font-size: 0.7rem;">{notif.created_at.strftime("%H:%M")} (Dibaca)</div>
                    </div>
                </div>
            </a>
        </div>
    </li>
    '''
    
    return HttpResponse(item_html + badge_html)


# =============================================================================
# EMPLOYEE MUTATION VIEW
# =============================================================================

@login_required
@require_POST
def mutate_employee_view(request, employee_id):
    """
    HTMX View untuk memproses mutasi karyawan.
    Only accessible by HRD or Superuser.
    """
    from attendance.models import EmployeeMutation
    
    # Permission Check: HRD or Superuser only
    if not request.user.is_superuser:
        if hasattr(request.user, 'employee_profile'):
            if request.user.employee_profile.role not in ['ADMIN', 'HRD']:
                return HttpResponse('<div class="alert alert-danger">Akses ditolak. Hanya HRD/Admin yang bisa memutasi karyawan.</div>', status=403)
        else:
            return HttpResponse('<div class="alert alert-danger">Akses ditolak.</div>', status=403)
    
    employee = get_object_or_404(Employee, id=employee_id)
    old_location = employee.home_base
    
    # Get form data
    new_location_id = request.POST.get('new_location')
    effective_date_str = request.POST.get('effective_date')
    reason = request.POST.get('reason', '').strip()
    
    if not new_location_id or not effective_date_str:
        return HttpResponse('<div class="alert alert-warning">Lokasi tujuan dan tanggal efektif wajib diisi.</div>')
    
    try:
        from datetime import datetime
        new_location = get_object_or_404(WorkLocation, id=new_location_id)
        effective_date = datetime.strptime(effective_date_str, '%Y-%m-%d').date()
        
        # Validate: Can't mutate to same location
        if old_location and old_location.id == new_location.id:
            return HttpResponse('<div class="alert alert-warning">Lokasi tujuan sama dengan lokasi asal.</div>')
        
        with transaction.atomic():
            # 1. Create mutation log
            EmployeeMutation.objects.create(
                employee=employee,
                old_location=old_location,
                new_location=new_location,
                effective_date=effective_date,
                reason=reason if reason else None,
                created_by=request.user
            )
            
            # 2. Update employee's home_base
            employee.home_base = new_location
            employee.save(update_fields=['home_base'])
        
        # Success response
        return HttpResponse(f'''
            <div class="alert alert-success">
                <i class="fas fa-check-circle me-2"></i>
                <strong>Mutasi Berhasil!</strong><br>
                {employee.full_name} dipindahkan dari <strong>{old_location.name if old_location else '-'}</strong> 
                ke <strong>{new_location.name}</strong> efektif tanggal {effective_date.strftime('%d %b %Y')}.
            </div>
            <script>
                // Close modal and refresh location detail
                setTimeout(() => {{
                    bootstrap.Modal.getInstance(document.getElementById('mutationModal')).hide();
                    // Trigger refresh on active location node
                    const activeNode = document.querySelector('.node-item.active');
                    if (activeNode) activeNode.click();
                }}, 1500);
            </script>
        ''')
        
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Error: {str(e)}</div>')


# =============================================================================
# EMPLOYEE DETAIL VIEW
# =============================================================================

@login_required
def employee_detail_view(request, employee_id):
    """
    HTMX View untuk menampilkan detail lengkap karyawan.
    Includes: Profile, Attendance Stats, Mutation History
    """
    from attendance.models import Employee, AttendanceLog, EmployeeMutation, EmployeeShiftAssignment
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    employee = get_object_or_404(Employee, id=employee_id)
    now = timezone.now()
    
    # === ATTENDANCE STATS (Current Month) ===
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    attendance_logs = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__gte=start_of_month
    )
    
    # Count by status
    stats = attendance_logs.values('status').annotate(count=Count('id'))
    stats_dict = {s['status']: s['count'] for s in stats}
    
    hadir = stats_dict.get('HADIR', 0)
    izin = stats_dict.get('IZIN', 0)
    sakit = stats_dict.get('SAKIT', 0)
    alpha = stats_dict.get('ALPHA', 0)
    
    # Working days in month (approximate)
    import calendar
    _, days_in_month = calendar.monthrange(now.year, now.month)
    working_days = days_in_month - 4  # Rough estimate excluding weekends
    
    # === RECENT LOGS (Last 10) ===
    recent_logs = AttendanceLog.objects.filter(employee=employee).order_by('-timestamp')[:10]
    
    # === MUTATION HISTORY ===
    mutations = EmployeeMutation.objects.filter(employee=employee).order_by('-effective_date')[:5]
    
    # === SHIFT ASSIGNMENT ===
    try:
        shift_assignment = EmployeeShiftAssignment.objects.filter(
            employee=employee,
            is_active=True
        ).select_related('shift_pattern').first()
    except:
        shift_assignment = None
    
    # === ATTENDANCE HEATMAP (Last 30 days) ===
    thirty_days_ago = now - timedelta(days=30)
    heatmap_logs = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__gte=thirty_days_ago,
        log_category='MASUK'
    ).values('timestamp__date').annotate(count=Count('id'))
    
    attendance_dates = {str(log['timestamp__date']): log['count'] for log in heatmap_logs}
    
    context = {
        'emp': employee,
        'stats': {
            'hadir': hadir,
            'izin': izin,
            'sakit': sakit,
            'alpha': alpha,
            'working_days': working_days,
            'hadir_pct': round((hadir / max(working_days, 1)) * 100),
        },
        'recent_logs': recent_logs,
        'mutations': mutations,
        'shift_assignment': shift_assignment,
        'attendance_dates': attendance_dates,
        'current_month': now.strftime('%B %Y'),
    }
    
    return render(request, 'portal/partials/_employee_detail.html', context)


# =============================================================================
# EMPLOYEE EDIT VIEW
# =============================================================================

@login_required
def employee_edit_view(request, employee_id):
    """
    View untuk mengedit data karyawan.
    GET: Render form edit dalam modal
    POST: Simpan perubahan data
    """
    from attendance.models import Employee, Department, WorkLocation, ShiftPattern, EmployeeShiftAssignment
    
    # Permission Check: Allow KERANI, HRD, ADMIN, SUPERUSER
    is_full_edit = False
    
    if request.user.is_superuser:
        is_full_edit = True
    elif hasattr(request.user, 'employee_profile'):
        role = request.user.employee_profile.role
        if role in ['ADMIN', 'HRD', 'SUPERVISOR']:
            is_full_edit = True
        elif role == 'KERANI':
            is_full_edit = False # Basic edit only
        else:
            return HttpResponse('<div class="alert alert-danger">Akses ditolak. Role tidak diizinkan.</div>', status=403)
    else:
        return HttpResponse('<div class="alert alert-danger">Akses ditolak.</div>', status=403)
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST':
        # Get form data common to all
        full_name = request.POST.get('full_name', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Validation
        if not full_name:
            return HttpResponse('<div class="alert alert-warning">Nama lengkap wajib diisi.</div>')
        
        try:
            with transaction.atomic():
                # Update Common Fields
                employee.full_name = full_name
                employee.is_active = is_active
                
                # Update Advanced Fields (Full Edit Only)
                if is_full_edit:
                    employee_type = request.POST.get('employee_type', '')
                    department_id = request.POST.get('department', '')
                    phone_number = request.POST.get('phone_number', '').strip()
                    telegram_user_id = request.POST.get('telegram_user_id', '').strip()
                    device_user_id = request.POST.get('device_user_id', '').strip()
                    joined_date_str = request.POST.get('joined_date', '')
                    shift_pattern_id = request.POST.get('shift_pattern', '')
                    attendance_method = request.POST.get('attendance_method', 'FINGERPRINT')

                    employee.employee_type = employee_type
                    employee.phone_number = phone_number if phone_number else None
                    employee.telegram_user_id = telegram_user_id if telegram_user_id else None
                    employee.attendance_method = attendance_method
                    
                    # Device User ID
                    if device_user_id:
                        employee.device_user_id = int(device_user_id)
                    else:
                        employee.device_user_id = None
                    
                    # Department
                    if department_id:
                        employee.department_id = department_id
                    else:
                        employee.department = None
                    
                    # Joined Date
                    if joined_date_str:
                        from datetime import datetime
                        employee.joined_date = datetime.strptime(joined_date_str, '%Y-%m-%d').date()
                    
                    # Handle Shift Assignment
                    if shift_pattern_id:
                        shift_pattern = get_object_or_404(ShiftPattern, id=shift_pattern_id)
                        # Deactivate previous assignments
                        EmployeeShiftAssignment.objects.filter(employee=employee, is_active=True).update(is_active=False)
                        # Create new assignment with effective_from date
                        from django.utils import timezone
                        EmployeeShiftAssignment.objects.create(
                            employee=employee,
                            shift_pattern=shift_pattern,
                            is_active=True,
                            effective_from=timezone.now().date()
                        )

                employee.save()
            
            # Success response - reload current employee detail
            return HttpResponse(f'''
                <div class="alert alert-success">
                    <i class="fas fa-check-circle me-2"></i>
                    Data <strong>{employee.full_name}</strong> berhasil diperbarui!
                </div>
                <script>
                    setTimeout(() => {{
                        // Close modal
                        bootstrap.Modal.getInstance(document.getElementById('editEmployeeModal')).hide();
                        // Reload employee detail
                        htmx.ajax('GET', '/portal/employee/{employee.id}/', {{target: '#location-content-area', swap: 'innerHTML'}});
                    }}, 1000);
                </script>
            ''')
            
        except Exception as e:
            return HttpResponse(f'<div class="alert alert-danger">Error: {str(e)}</div>')
    
    # GET: Render form
    departments = Department.objects.all()
    
    # RBAC: Filter Departments for Supervisor & Kerani
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        if profile.role in ['SUPERVISOR', 'KERANI'] and profile.assigned_location:
            # Get location scope (include descendants)
            user_locations = profile.assigned_location.get_descendants(include_self=True)
            
            # Determine Location Type Compatibility (Chicken & Egg Fix)
            # Use keywords in Location Name/Code to guess type
            loc_types = []
            
            # Check root assigned location
            root_loc = profile.assigned_location
            name_upper = root_loc.name.upper()
            code_upper = root_loc.code.upper()
            
            if code_upper == 'HO' or 'HEAD OFFICE' in name_upper:
                loc_types.append('HO')
            elif 'PKS' in name_upper or 'MILL' in name_upper or 'PABRIK' in name_upper or 'PKS' in code_upper:
                loc_types.append('MILL')
            else:
                # Default to ESTATE for Kebun, Afdeling, etc.
                loc_types.append('ESTATE')
            
            # Filter departments:
            # 1. Have active employees in my territory OR
            # 2. Match my location type (even if empty)
            from django.db.models import Q
            departments = departments.filter(
                Q(employees__home_base__in=user_locations, employees__is_active=True) |
                Q(location_type__in=loc_types)
            ).distinct()

    shift_patterns = ShiftPattern.objects.all()
    
    # Get current shift assignment
    current_shift = None
    try:
        current_shift = EmployeeShiftAssignment.objects.filter(
            employee=employee, is_active=True
        ).select_related('shift_pattern').first()
    except:
        pass
    
    context = {
        'emp': employee,
        'departments': departments,
        'shift_patterns': shift_patterns,
        'current_shift': current_shift,
        'employee_types': Employee.TYPE_CHOICES,
        'attendance_methods': Employee.METHOD_CHOICES,
        'is_full_edit': is_full_edit,
    }
    
    return render(request, 'portal/partials/_employee_edit_form.html', context)


# =============================================================================
# RECAP MATRIX VIEW (Meja Kerja Rekapitulasi)
# =============================================================================

@login_required
@never_cache
def recap_matrix_view(request):
    """
    Meja Kerja Rekapitulasi - Matriks Absensi Bulanan
    
    RBAC Rules:
    - KERANI: Hanya melihat lokasi sendiri (assigned_location)
    - HRD/ADMIN: Bisa memilih lokasi apapun
    """
    from attendance.models import WorkLocation, Employee, AttendanceLog
    from django.utils import timezone
    from django.db.models import F
    from calendar import monthrange
    from datetime import date, timedelta
    from collections import defaultdict
    import holidays  # Holiday detection (Reload Triggered)
    
    # Get current date for defaults
    now = timezone.now()
    
    # Get filter parameters
    try:
        month = int(request.GET.get('month', now.month))
        year = int(request.GET.get('year', now.year))
    except (ValueError, TypeError):
        month = now.month
        year = now.year
    
    # Validate month range
    if month < 1 or month > 12:
        month = now.month
    
    # =========================================================================
    # RBAC: STRICT LOCATION FILTERING
    # =========================================================================
    selected_location = None
    all_locations = []
    user_role = None
    can_select_location = False
    profile = None  # Initialize profile to prevent UnboundLocalError
    
    if hasattr(request.user, 'employee_profile'):
        profile = request.user.employee_profile
        user_role = profile.role
        
        if profile.role in ['KERANI', 'SUPERVISOR', 'DEPT_ADMIN']:
            # LOCATION RESTRICTED ROLES: Can see assigned location AND its descendants
            if profile.assigned_location:
                # Get descendants including self
                user_locations = profile.assigned_location.get_descendants(include_self=True)
                
                # Filter only leaf nodes for the dropdown logic if that's the desired behavior, 
                # OR allow selecting any node in their subtree.
                # The prompt asks for "Maro Sebo" (Parent) AND "Afdeling 6" (Child).
                # So we should probably allow all nodes in the subtree, or at least consistent with other views.
                
                # Let's populate all_locations with the subtree
                all_locations = user_locations.order_by('tree_id', 'lft')
                
                # Allow selection if there's more than 1 location in the subtree
                # For DEPT_ADMIN, they might be restricted to just one location if needed, but per plan they follow location scope
                if all_locations.count() > 1:
                    can_select_location = True
                else:
                    can_select_location = False
                
                # Handle selection
                location_id = request.GET.get('location_id')
                if location_id:
                    try:
                        # Ensure selected location is within their scope
                        target_loc = WorkLocation.objects.get(id=location_id)
                        if target_loc in user_locations:
                            selected_location = target_loc
                        else:
                            # Fallback if tried to access unauthorized location
                            selected_location = profile.assigned_location
                    except WorkLocation.DoesNotExist:
                        selected_location = profile.assigned_location
                else:
                     # Default to assigned location (Parent)
                    selected_location = profile.assigned_location
            else:
                 # Fallback if no assigned location
                 pass

        if profile.role in ['ADMIN', 'HRD', 'ACCOUNTING']:
            can_select_location = True
            # Allow selecting ANY location (not just leaf nodes)
            all_locations = WorkLocation.objects.all().order_by('tree_id', 'lft')
            
            # Allow filter override
            location_id = request.GET.get('location_id')
            if location_id:
                try:
                    selected_location = WorkLocation.objects.get(id=location_id)
                except WorkLocation.DoesNotExist:
                    pass
            
            # Default to first location if none selected (or if assigned_location is not in list)
            if not selected_location and all_locations.exists():
                selected_location = all_locations.first()
    else:
        # Fallback for superuser/staff without employee_profile
        if request.user.is_superuser or request.user.is_staff:
            can_select_location = True
            # Allow selecting ANY location
            all_locations = WorkLocation.objects.all().order_by('tree_id', 'lft')
            
            location_id = request.GET.get('location_id')
            if location_id:
                try:
                    selected_location = WorkLocation.objects.get(id=location_id)
                except WorkLocation.DoesNotExist:
                    pass
            if not selected_location and all_locations.exists():
                selected_location = all_locations.first()

    # --- DEBUG LOCATION SELECTION ---
    try:
        debug_log_path = 'd:\\dev\\attendance_core\\debug_location.txt'
        with open(debug_log_path, 'a') as f:
            f.write(f"\n[{timezone.now()}] Location Selection Debug:\n")
            f.write(f"User: {request.user} (Superuser: {request.user.is_superuser})\n")
            if hasattr(request.user, 'employee_profile'):
                f.write(f"Profile: {request.user.employee_profile}\n")
                f.write(f"Assigned Location: {request.user.employee_profile.assigned_location}\n")
            else:
                f.write("Profile: None\n")
            f.write(f"Selected Location: {selected_location}\n")
    except Exception as e:
        print(f"Log Error: {e}")
    # --------------------------------
    # =========================================================================
    # MONTHLY REPORT WORKFLOW (Draft -> Publish -> Locked)
    # =========================================================================
    from attendance.models import MonthlyReport
    from django.utils import timezone
    
    report = None
    is_locked = False
    error_message = "" # Initialize variable
    
    if selected_location:
        # === FIX: Pastikan Parameter adalah Integer ===
        # Ini mencegah mismatch query jika 'month' dari URL berbentuk string
        try:
            month_int = int(month)
            year_int = int(year)
        except (ValueError, TypeError):
            month_int = timezone.now().month
            year_int = timezone.now().year

        # === LOGIC UTAMA: AUTO-GET OR CREATE REPORT ===
        # Coba ambil report yang ada, atau buat baru jika parameter tidak cocok
        try:
            # === LOGIC UTAMA: AUTO-GET OR CREATE REPORT ===
            try:
                report, created = MonthlyReport.objects.get_or_create(
                    location=selected_location,
                    period_month=month_int,
                    period_year=year_int,
                    defaults={
                        'status': 'DRAFT',
                        'version': 0
                    }
                )
                
                # if created:
                #      print(f"[DEBUG] MonthlyReport Created: {report}")
                # else:
                #      print(f"[DEBUG] MonthlyReport Fetched: {report}")

            except Exception as e:
                # If get_or_create fails (unlikely, maybe unexpected kwargs), capture it
                err_msg = f"get_or_create failed: {str(e)}"
                print(f"[ERROR] {err_msg}")
                raise e # Propagate to outer try/except for UI handling
            
            # === Tentukan Status Lock ===
            # Lock edit jika status bukan DRAFT dan bukan REQ_REVISI
            is_locked = report.status not in ['DRAFT', 'REQ_REVISI']
            
        except Exception as e:
            err_msg = f"Failed to get/create report: {str(e)}"
            import traceback
            error_message = f"{str(e)} | Truncated Trace: {traceback.format_exc()[:300]}" # Capture for UI
            
            print(f"[ERROR] {err_msg}")
            
            # Show error to user in UI
            from django.contrib import messages
            messages.error(request, err_msg)
            
            traceback.print_exc()
            report = None
            is_locked = False
            error_message = str(e) # Pass to context
    else:
        # Fallback if no location selected
        from django.contrib import messages
        messages.warning(request, "Lokasi kerja tidak ditemukan atau belum dipilih.")
        print("[DEBUG] selected_location is None, skipping MonthlyReport")
        error_message = "No Location Selected"
    
    # =========================================================================
    # VIEW MODE SELECTION
    # =========================================================================
    employee_id = request.GET.get('employee_id')
    employee_search = request.GET.get('employee_search', '').strip()
    wa_employee_search = request.GET.get('wa_employee_search', '').strip()
    mode = 'matrix'
    personal_logs = []
    summary_stats = {'late': 0, 'overtime': 0, 'alpha': 0, 'permit': 0, 'sick': 0, 'hadir': 0, 'late_count': 0}
    personal_employee = None  # Will hold target employee in personal mode
    
    # If specific employee selected, switch to PERSONAL MODE
    if employee_id:
        try:
            target_employee = Employee.objects.get(id=employee_id)
            personal_employee = target_employee  # Set for template context
            mode = 'personal'
            
            # Build date range for the month
            _, days_in_month = monthrange(year, month)
            date_range = [date(year, month, day) for day in range(1, days_in_month + 1)]
            
            # Fetch logs for this employee only
            logs = AttendanceLog.objects.filter(
                employee=target_employee,
                timestamp__year=year,
                timestamp__month=month
            ).order_by('timestamp')
            
            # Group logs by LOCAL date (crucial for correct daily bucket)
            daily_logs = defaultdict(list)
            for log in logs:
                # Convert to local time first!
                local_dt = timezone.localtime(log.timestamp)
                daily_logs[local_dt.date()].append({
                    'log': log,
                    'local_dt': local_dt
                })
            
            # Standard Schedule (Hardcoded for now, can be moved to model later)
            # SCHEDULE_IN = timezone.datetime.strptime('08:00', '%H:%M').time()
            # SCHEDULE_OUT = timezone.datetime.strptime('17:00', '%H:%M').time()
            
            # Holiday Object for Indonesia
            id_holidays = holidays.ID(years=[year, year-1])
            
            for d in date_range:
                # Get records for this local date
                day_records = daily_logs.get(d, [])
                
                day_stat = {
                    'date': d,
                    'is_weekend': d.weekday() == 6,  # Only Sunday is weekend (holiday)
                    'is_holiday': d in id_holidays,
                    'holiday_name': id_holidays.get(d),
                    'schedule_in': '08:00',
                    'schedule_out': '17:00',
                    'clock_in': '-',
                    'clock_out': '-',
                    'late_minutes': 0,
                    'overtime_minutes': 0,
                    'status': '',
                    'status_class': '',
                    'raw_status': '',
                    # 5 Checkpoint times
                    'cp_masuk': None,
                    'cp_1': None,
                    'cp_istirahat': None,
                    'cp_2': None,
                    'cp_pulang': None,
                    # Count filled checkpoints
                    'checkpoints_filled': 0,
                }
                
                if day_records:
                    # ... (existing logic) ...
                    # Sort by time
                    day_records.sort(key=lambda x: x['local_dt'])
                    
                    # Extract checkpoints from logs based on log_category
                    for rec in day_records:
                        log = rec['log']
                        local_dt = rec['local_dt']
                        time_str = local_dt.strftime('%H:%M')
                        
                        # Map log_category to checkpoint fields
                        cat = getattr(log, 'log_category', None)
                        if cat == 'MASUK':
                            # Capture EARLIEST Masuk for display
                            if not day_stat['cp_masuk']:
                                day_stat['cp_masuk'] = time_str
                            day_stat['checkpoints_filled'] += 1
                        elif cat == 'CP_1':
                            day_stat['cp_1'] = time_str
                            day_stat['checkpoints_filled'] += 1
                        elif cat == 'ISTIRAHAT':
                            day_stat['cp_istirahat'] = time_str
                            day_stat['checkpoints_filled'] += 1
                        elif cat == 'CP_2':
                            day_stat['cp_2'] = time_str
                            day_stat['checkpoints_filled'] += 1
                        elif cat == 'PULANG':
                            day_stat['cp_pulang'] = time_str
                            day_stat['checkpoints_filled'] += 1
                    
                    # Fallback: If no log_category, use first/last for clock_in/out
                    first_record = day_records[0]
                    first_log = first_record['log']
                    local_in = first_record['local_dt']
                    
                    day_stat['clock_in'] = local_in.strftime('%H:%M')
                    day_stat['status'] = first_log.get_status_display() if hasattr(first_log, 'get_status_display') else ''
                    day_stat['raw_status'] = getattr(first_log, 'status', '')
                    
                    # If cp_masuk not set from category, use first log
                    if not day_stat['cp_masuk']:
                        day_stat['cp_masuk'] = day_stat['clock_in']
                        day_stat['checkpoints_filled'] += 1
                    
                    # Calculate Late
                    cin_time = local_in.time()
                    dummy_date = date(2000, 1, 1)
                    
                    # Dynamic Schedule & Late Tolerance
                    daily_sch, shift_pattern = get_employee_schedule(personal_employee, d)
                    
                    # Default values
                    sch_in = DEFAULT_SCHEDULE_IN
                    tol = 0
                    
                    if daily_sch:
                        sch_in = daily_sch.clock_in
                        tol = daily_sch.late_tolerance
                    
                    # Threshold = Scheduled In + Tolerance
                    dt_sch = datetime.combine(dummy_date, sch_in)
                    dt_threshold = dt_sch + timedelta(minutes=tol)
                    
                    # Check if Late (Actual Time > Threshold)
                    if cin_time > dt_threshold.time() and not day_stat['is_holiday'] and not day_stat['is_weekend']:
                        dt_in = datetime.combine(dummy_date, cin_time)
                        # Late calculated from Threshold Time (tolerance not counted as late)
                        diff = (dt_in - dt_threshold).total_seconds() / 60
                        
                        mins_late = int(diff)
                        day_stat['late_minutes'] = mins_late
                        summary_stats['late'] += mins_late
                        
                        if mins_late > 0:
                            summary_stats['late_count'] += 1
                    
                    # Ensure 'HADIR' status also counts as Hadir
                    if first_log.status == 'HADIR' and not day_stat['status_class']:
                         summary_stats['hadir'] += 1
                    
                    # Clock Out (Last Record if > 1)
                    if len(day_records) > 1:
                        last_record = day_records[-1]
                        local_out = last_record['local_dt']
                        day_stat['clock_out'] = local_out.strftime('%H:%M')
                        
                        # If cp_pulang not set from category, use last log
                        if not day_stat['cp_pulang']:
                            day_stat['cp_pulang'] = day_stat['clock_out']
                            day_stat['checkpoints_filled'] += 1
                        
                        # Calculate Overtime
                        cout_time = local_out.time()
                        
                        # Dynamic Schedule Out with Overtime Tolerance
                        sch_out = DEFAULT_SCHEDULE_OUT
                        ot_tol = 0  # overtime tolerance in minutes
                        if daily_sch:
                            sch_out = daily_sch.clock_out
                            ot_tol = daily_sch.overtime_tolerance
                        
                        # Calculate overtime threshold (clock_out + tolerance)
                        dt_sch_out = datetime.combine(dummy_date, sch_out)
                        dt_ot_threshold = dt_sch_out + timedelta(minutes=ot_tol)
                        
                        # Overtime only counts if employee leaves AFTER the threshold
                        if cout_time > dt_ot_threshold.time():
                            dt_out = datetime.combine(dummy_date, cout_time)
                            # Overtime calculated from threshold time, not clock_out
                            diff = (dt_out - dt_ot_threshold).total_seconds() / 60
                            day_stat['overtime_minutes'] = int(diff)
                            summary_stats['overtime'] += int(diff)
                    
                    # Determine Status Class
                    if first_log.status == 'CHECKIN':
                        day_stat['status_class'] = 'text-success fw-bold'
                        summary_stats['hadir'] += 1
                        if day_stat['is_holiday']:
                             day_stat['status'] += ' (Hari Libur)'
                    elif first_log.status == 'SICK':
                        day_stat['status_class'] = 'text-warning fw-bold'
                        summary_stats['sick'] += 1
                    elif first_log.status == 'PERMIT':
                        day_stat['status_class'] = 'text-info fw-bold'
                        summary_stats['permit'] += 1
                    elif first_log.status == 'ALPHA':
                        day_stat['status_class'] = 'text-danger fw-bold'
                        summary_stats['alpha'] += 1
                        
                else:
                    # No logs
                    # Check for Leave FIRST (Priority over Holiday/Weekend/Alpha)
                    leave = EmployeeLeave.objects.filter(
                        employee=personal_employee,
                        start_date__lte=d,
                        end_date__gte=d
                    ).first()
                    
                    if leave:
                        day_stat['status'] = leave.get_leave_type_display()
                        if leave.leave_type == 'SAKIT':
                            day_stat['status_class'] = 'text-warning fw-bold'
                            summary_stats['sick'] += 1
                        elif leave.leave_type == 'IZIN':
                            day_stat['status_class'] = 'text-info fw-bold'
                            summary_stats['permit'] += 1
                        elif leave.leave_type == 'CUTI':
                            day_stat['status_class'] = 'text-success fw-bold'
                            # summary_stats['leave'] += 1 
                        else:
                            day_stat['status_class'] = 'text-secondary'
                            
                    elif day_stat['is_holiday']:
                        day_stat['status'] = 'Libur: ' + (day_stat['holiday_name'] or '')
                        day_stat['status_class'] = 'text-secondary'
                        
                    elif day_stat['is_weekend']:
                        day_stat['status'] = 'Minggu'
                        day_stat['status_class'] = 'text-secondary'

                    elif d < now.date():
                        day_stat['status'] = 'Alpha'
                        day_stat['status_class'] = 'text-danger fw-bold'
                        summary_stats['alpha'] += 1
                        
                    else:
                        day_stat['status'] = '-'
                
                personal_logs.append(day_stat)
                
        except Employee.DoesNotExist:
            mode = 'matrix'
            
    
    # =========================================================================
    # BUILD MATRIX DATA (Only if Mode == MATRIX)
    # =========================================================================
    employees = []
    matrix_data = {}
    date_range = []
    fp_summary_stats = {}
    wa_summary_data = {}
    
    # FETCH DEPARTMENTS FOR FILTER
    from attendance.models import Department
    all_departments = Department.objects.all()
    selected_department_id = request.GET.get('department_id')
    
    if selected_location:
        # Get all sub-locations (include self)
        sub_locations = selected_location.get_descendants(include_self=True)
        
        # Get active employees in these locations
        employees_qs = Employee.objects.filter(
            home_base__in=sub_locations,
            is_verified=True,
            is_active=True
        )
        
        # FILTER BY DEPARTMENT IF SELECTED
        if selected_department_id:
            employees_qs = employees_qs.filter(department_id=selected_department_id)
        
        # SUPERVISOR / DEPT_ADMIN LOGIC
        if profile and profile.role == 'DEPT_ADMIN':
            if profile.assigned_department:
                employees_qs = employees_qs.filter(department=profile.assigned_department)
                # Filter dropdown options to only the assigned department
                all_departments = all_departments.filter(id=profile.assigned_department.id)

        # SUPERVISOR & KERANI LOGIC: Filter departments to those active in their location
        if profile and profile.role in ['SUPERVISOR', 'KERANI'] and profile.assigned_location:
            # Get location scope (include descendants)
            sup_locations = profile.assigned_location.get_descendants(include_self=True)
            # Filter departments that have active employees in these locations
            all_departments = all_departments.filter(
                employees__home_base__in=sup_locations,
                employees__is_active=True
            ).distinct()
            
        employees = employees_qs.select_related('department', 'home_base').order_by('full_name')
        
        page_obj = None
        if mode == 'matrix':
            # --- TAB AWARE PAGINATION ---
            # Determine active tab
            active_tab = request.GET.get('active_tab', 'fingerprint')
            
            # 1. Split QuerySets for Counts (Global for Location)
            # Use Q objects to ensure coverage. Manual defaults to Fingerprint.
            fp_qs = employees.filter(Q(attendance_method='FINGERPRINT') | Q(attendance_method='MANUAL')).exclude(attendance_method='WHATSAPP')
            wa_qs = employees.filter(attendance_method='WHATSAPP')
            
            # --- SEARCH FILTERING ---
            employee_search = request.GET.get('employee_search', '').strip()
            wa_employee_search = request.GET.get('wa_employee_search', '').strip()
            
            if employee_search:
                fp_qs = fp_qs.filter(Q(full_name__icontains=employee_search) | Q(nik__icontains=employee_search))
                
            if wa_employee_search:
                wa_qs = wa_qs.filter(Q(full_name__icontains=wa_employee_search) | Q(nik__icontains=wa_employee_search))
            
            # ------------------------
            
            fp_count = fp_qs.count()
            wa_count = wa_qs.count()
            
            # 2. Select Target QuerySet based on Active Tab
            if active_tab == 'whatsapp':
                target_qs = wa_qs
            else:
                target_qs = fp_qs

            # Build date range for the month
            _, days_in_month = monthrange(year, month)
            date_range = [date(year, month, day) for day in range(1, days_in_month + 1)]
            
            # --- OPTIMIZATION 1: PAGINATION ---
            from django.core.paginator import Paginator
            
            # Get per_page from request, default to 20
            try:
                per_page = int(request.GET.get('per_page', 20))
                if per_page not in [10, 20, 50, 100]:
                    per_page = 20
            except ValueError:
                per_page = 20
            
            paginator = Paginator(target_qs, per_page)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            employees = page_obj.object_list # Slicing QuerySet

            # --- OPTIMIZATION 2: BULK PREFETCH SHIFTS ---
            from attendance.models import EmployeeShiftAssignment
            # Fetch all active assignments for these specific employees
            # We fetch assignments that overlap with the current month or are open-ended
            month_start = date(year, month, 1)
            month_end = date(year, month, days_in_month)
            
            bulk_assignments = EmployeeShiftAssignment.objects.filter(
                employee__in=employees,
                is_active=True
            ).filter(
                Q(effective_to__isnull=True) | Q(effective_to__gte=month_start)
            ).filter(
                effective_from__lte=month_end
            ).select_related('shift_pattern').order_by('employee', '-effective_from')

            # Map assignments to employee_id
            assignment_map = defaultdict(list)
            for asm in bulk_assignments:
                assignment_map[asm.employee_id].append(asm)

            # Fetch logs for PAGINATED employees only
            logs = AttendanceLog.objects.filter(
                employee__in=employees,
                timestamp__year=year,
                timestamp__month=month
            ).select_related('employee').order_by('employee', 'timestamp')
            
            # Group logs by employee and date
            employee_date_logs = defaultdict(lambda: defaultdict(list))
            for log in logs:
                # Convert to local time to ensure correct date bucket
                local_dt = timezone.localtime(log.timestamp)
                employee_date_logs[log.employee_id][local_dt.date()].append(log)
            
            # Build matrix: { employee_id: { date_obj: status_code } }
            
            # Holiday Object for Indonesia
            id_holidays = holidays.ID(years=[year, year-1])

            # Fetch leaves for PAGINATED employees
            leaves = EmployeeLeave.objects.filter(
                employee__in=employees,
                start_date__lte=date_range[-1],
                end_date__gte=date_range[0]
            )
            
            # Map leaves to (emp_id, date) -> leave_type
            employee_leaves_map = defaultdict(lambda: defaultdict(str))
            for leave in leaves:
                # Iterate dates for this leave
                curr = max(leave.start_date, date_range[0])
                end = min(leave.end_date, date_range[-1])
                while curr <= end:
                    employee_leaves_map[leave.employee_id][curr] = leave.leave_type
                    curr += timedelta(days=1)

            for emp in employees:
                emp_matrix = {}
                emp_stats = {'H': 0, 'A': 0, 'T': 0, 'S': 0, 'I': 0}
                
                # Get employee's assignments from memory
                emp_assignments = assignment_map.get(emp.id, [])

                for d in date_range:
                    day_logs = employee_date_logs.get(emp.id, {}).get(d, [])
                    is_holiday = d in id_holidays
                    
                    if day_logs:
                        # Check specific status from log
                        first_log = day_logs[0]
                        status_map = {
                            'CHECKIN': 'H', 'SICK': 'S', 
                            'PERMIT': 'I', 'ALPHA': 'A'
                        }
                        st = status_map.get(first_log.status, 'H')
                        emp_matrix[d] = st
                        
                        # Stats Calculation
                        if st == 'H':
                            emp_stats['H'] += 1
                            # Check Late
                            local_dt = timezone.localtime(first_log.timestamp)
                            
                            # --- OPTIMIZED SCHEDULE RETRIEVAL ---
                            # Find applicable assignment in memory
                            daily_sch = None
                            for asm in emp_assignments:
                                if asm.effective_from <= d and (asm.effective_to is None or asm.effective_to >= d):
                                    if asm.shift_pattern:
                                        daily_sch = asm.shift_pattern.get_schedule_for_day(d.weekday())
                                    break
                            
                            sch_in = DEFAULT_SCHEDULE_IN
                            tol = 0
                            if daily_sch:
                                sch_in = daily_sch.clock_in
                                tol = daily_sch.late_tolerance
                                
                            dummy_date = date(2000, 1, 1)
                            dt_sch = datetime.combine(dummy_date, sch_in)
                            dt_threshold = dt_sch + timedelta(minutes=tol)
                            
                            # Only count late if NOT holiday and NOT Sunday
                            is_weekend = d.weekday() == 6  # Sunday only
                            
                            # Round clock-in time to minutes (ignore seconds)
                            clock_time_rounded = local_dt.time().replace(second=0, microsecond=0)
                            threshold_time_rounded = dt_threshold.time().replace(second=0, microsecond=0)
                            
                            if clock_time_rounded > threshold_time_rounded and not is_holiday and not is_weekend:
                                emp_stats['T'] += 1
                        elif st == 'S':
                            emp_stats['S'] += 1
                        elif st == 'I':
                            emp_stats['I'] += 1
                        elif st == 'A':
                            emp_stats['A'] += 1

                    else:
                        # No logs - Check LEAVE first
                        l_type = employee_leaves_map[emp.id].get(d)
                        
                        if l_type:
                            code_map = {'IZIN': 'I', 'SAKIT': 'S', 'CUTI': 'C', 'CUTI_KHUSUS': 'C'}
                            st = code_map.get(l_type, 'I')
                            emp_matrix[d] = st
                            # Map Cuti to Izin stats or leave it?
                            # Existing stats only has H, A, T, S, I. Map C -> I or ignore?
                            # Let's map C -> I for now as per minimal stats
                            stat_key = 'I' if st == 'C' else st
                            if stat_key in emp_stats:
                                emp_stats[stat_key] += 1
                                
                        elif is_holiday:
                            emp_matrix[d] = 'L' # Libur Nasional
                        elif d.weekday() == 6: # Sunday Only
                            emp_matrix[d] = 'L' # Libur Weekend
                        elif d < now.date():
                            emp_matrix[d] = 'A'
                            emp_stats['A'] += 1
                        else:
                            emp_matrix[d] = ''
                
                matrix_data[emp.id] = emp_matrix
                fp_summary_stats[emp.id] = emp_stats
            
    # Prepare Holidays Map for Template (Date String -> Name)
    # MOVED OUTSIDE if mode == 'matrix' block to fix UnboundLocalError
    holidays_map = {}
    id_holidays = holidays.ID(years=[year, year-1]) 
    
    for d in date_range:
        if d in id_holidays:
            # Key must be string to match template date filter
            holidays_map[d.strftime('%Y-%m-%d')] = id_holidays.get(d)
        elif d.weekday() == 6:
            # Explicitly mark Sundays for template coloring
            holidays_map[d.strftime('%Y-%m-%d')] = "Minggu"

    # =========================================================================
    # WHATSAPP / TELEGRAM 5-POINT TRACKING
    # =========================================================================
    # Parse specific date for WA Daily View (default: today)
    wa_specific_date_str = request.GET.get('wa_date', now.date().isoformat())
    try:
        wa_specific_date = date.fromisoformat(wa_specific_date_str)
    except ValueError:
        wa_specific_date = now.date()
    
    # Split employees by source type
    fingerprint_employees = []
    wa_employees = []

    if employees: # Use filtered 'employees' list, but we need ORIGINAL list for categorization if pagination didn't happen yet?
                  # Actually 'employees' variable above is already Paginated if mode=matrix
                  # But here we probably want the Full list for WA splits if we weren't in matrix mode?
                  # Wait, the logic below separates employees into FP and WA.
                  # If we use the PAGINATED 'employees' list, we only get the current page's items.
                  # That seems correct for "Visualizing what's on screen", but the WA tab might need its own pagination if it uses distinct logic?
                  # The code uses 'employees' which IS paginated if mode==matrix.
                  # If active_tab is WA, 'employees' contains WA employees (page N).
                  # If active_tab is FP, 'employees' contains FP employees (page N).
                  # So we iterate 'employees' and split them, meaning one list will be populated and the other empty/small.
        
        # NOTE: logic below relies on 'employees' being a list of Employee objects.
        
        # Get employees with confirmed log sources
        # Optim: IDs only
        emp_ids = [e.id for e in employees]
        
        has_wa_logs = set(AttendanceLog.objects.filter(
            employee_id__in=emp_ids,
            source_type='TELEGRAM'
        ).values_list('employee_id', flat=True))
        
        has_fp_logs = set(AttendanceLog.objects.filter(
            employee_id__in=emp_ids,
            source_type='FINGERPRINT'
        ).values_list('employee_id', flat=True))
        
        for e in employees:
            # 1. Explicit Attendance Method (Preferred)
            if e.attendance_method == 'FINGERPRINT':
                fingerprint_employees.append(e)
                continue
            elif e.attendance_method == 'WHATSAPP':
                wa_employees.append(e)
                continue
            
            # 2. Fallback (Legacy Heuristics - for safety or 'MANUAL')
            if e.id in has_wa_logs:
                wa_employees.append(e)
            elif e.id in has_fp_logs:
                fingerprint_employees.append(e)
            else:
                if e.device_user_id:
                     fingerprint_employees.append(e)
                elif e.employee_type == 'HARIAN':
                    wa_employees.append(e)
                else:
                    fingerprint_employees.append(e)
    
    # WA 5-Point Categories
    WA_CATEGORIES = ['MASUK', 'CHECKPOINT_1', 'ISTIRAHAT', 'CHECKPOINT_2', 'PULANG']
    WA_CATEGORY_ICONS = {
        'MASUK': ('fa-sun', 'Pagi'),
        'CHECKPOINT_1': ('fa-flag', 'CP1'),
        'ISTIRAHAT': ('fa-mug-hot', 'Istirahat'),
        'CHECKPOINT_2': ('fa-flag-checkered', 'CP2'),
        'PULANG': ('fa-home', 'Pulang'),
    }
    
    # WA Daily Data (Global View - All WA Employees for specific date)
    wa_daily_data = {}
    wa_employee_id = request.GET.get('wa_employee_id')
    wa_mode = 'matrix' if not wa_employee_id else 'personal'
    wa_personal_logs = []
    wa_personal_employee = None  # Will hold target WA employee in personal mode
    
    # WA MATRIX DATA: { emp_id: { date: { category: time_str } } }
    wa_matrix_data = {}
    wa_summary_data = {}
    
    # Build date range for WA matrix (same as fingerprint)
    _, wa_days_in_month = monthrange(year, month)
    wa_date_range = [date(year, month, day) for day in range(1, wa_days_in_month + 1)]

    if wa_employees:
        # Build set of holiday dates for quick lookup
        holidays_set = set(holidays_map.keys())
        
        # Initialize matrix for all employees
        for emp in wa_employees:
            wa_matrix_data[emp.id] = {}
            for d in wa_date_range:
                wa_matrix_data[emp.id][d] = {cat: None for cat in WA_CATEGORIES}
                wa_matrix_data[emp.id][d]['IZIN'] = None  # Track izin status
        
        # Fetch ALL logs for WA employees in this month
        wa_month_logs = AttendanceLog.objects.filter(
            employee__in=wa_employees,
            timestamp__year=year,
            timestamp__month=month,
            source_type='TELEGRAM'
        ).select_related('employee')
        
        # Fill in the matrix
        for log in wa_month_logs:
            local_time = timezone.localtime(log.timestamp)
            log_date = local_time.date()
            cat = log.log_category if log.log_category in WA_CATEGORIES else 'UNKNOWN'
            if cat in WA_CATEGORIES and log_date in wa_matrix_data.get(log.employee_id, {}):
                wa_matrix_data[log.employee_id][log_date][cat] = local_time.strftime('%H:%M')
            # Check for izin/sakit (if stored in notes or specific field)
            if hasattr(log, 'notes') and log.notes and ('IZIN' in log.notes.upper() or 'SAKIT' in log.notes.upper()):
                wa_matrix_data[log.employee_id][log_date]['IZIN'] = log.notes[:10]
        
        # Fetch leaves for all WA employees in range (Inclusive)
        wa_leaves = EmployeeLeave.objects.filter(
            employee__in=wa_employees,
            start_date__lte=wa_date_range[-1],
            end_date__gte=wa_date_range[0]
        )
        
        for leave in wa_leaves:
            # Determine intersection with month
            start = max(leave.start_date, wa_date_range[0])
            end = min(leave.end_date, wa_date_range[-1])
            
            curr = start
            while curr <= end:
                # Ensure emp and date exist in matrix
                if leave.employee_id in wa_matrix_data and curr in wa_matrix_data[leave.employee_id]:
                     status_text = leave.get_leave_type_display()
                     
                     # 1. Set IZIN field (for Summary Calculation)
                     wa_matrix_data[leave.employee_id][curr]['IZIN'] = status_text
                     
                     # 2. Update Display Cells if Empty (Visualize in Big Table)
                     # Only if NO LOG exists strictly (check MASUK is None)
                     if not wa_matrix_data[leave.employee_id][curr]['MASUK']:
                         for cat in WA_CATEGORIES:
                             wa_matrix_data[leave.employee_id][curr][cat] = status_text
                
                curr += timedelta(days=1)
        
        # Calculate summary per employee
        wa_summary_data = {}
        # We'll use a new dict for the 5-column stats to avoid breaking existing logic if needed
        # But actually let's just add to wa_summary_data
        
        for emp in wa_employees:
            wa_summary_data[emp.id] = {
                'missing_pagi': 0, 'missing_cp1': 0, 'missing_istirahat': 0,
                'missing_cp2': 0, 'missing_pulang': 0, 'izin_count': 0,
                'total_missing': 0,
                # New Stats
                'H': 0, 'A': 0, 'T': 0, 'S': 0, 'I': 0
            }
            for d in wa_date_range:
                date_str = d.isoformat()
                day_data = wa_matrix_data[emp.id][d]
                
                # Check Izin/Sakit first
                izin_status = day_data.get('IZIN')
                if izin_status:
                    if 'SAKIT' in izin_status.upper():
                        wa_summary_data[emp.id]['S'] += 1
                    else:
                        wa_summary_data[emp.id]['I'] += 1
                    wa_summary_data[emp.id]['izin_count'] += 1
                    # Don't count H/A/Missing if Izin
                    continue
                
                # Check Hadir (Masuk)
                masuk_time = day_data.get('MASUK')
                if masuk_time:
                    wa_summary_data[emp.id]['H'] += 1
                    # Check Late
                    # masuk_time is string "HH:MM"
                    if masuk_time > "08:00":
                        wa_summary_data[emp.id]['T'] += 1
                else:
                    # No Masuk
                    # Check Alpha: Not holiday, Not Sunday, Past Date
                    if d < now.date() and date_str not in holidays_set and d.weekday() != 6:
                         wa_summary_data[emp.id]['A'] += 1
                
                # Skip stats for future/holidays for Missing Point Calculation
                if d > now.date() or date_str in holidays_set or d.weekday() == 6:
                    continue

                if not day_data.get('MASUK'): wa_summary_data[emp.id]['missing_pagi'] += 1
                if not day_data.get('CHECKPOINT_1'): wa_summary_data[emp.id]['missing_cp1'] += 1
                if not day_data.get('ISTIRAHAT'): wa_summary_data[emp.id]['missing_istirahat'] += 1
                if not day_data.get('CHECKPOINT_2'): wa_summary_data[emp.id]['missing_cp2'] += 1
                if not day_data.get('PULANG'): wa_summary_data[emp.id]['missing_pulang'] += 1
            
            wa_summary_data[emp.id]['total_missing'] = (
                wa_summary_data[emp.id]['missing_pagi'] +
                wa_summary_data[emp.id]['missing_cp1'] +
                wa_summary_data[emp.id]['missing_istirahat'] +
                wa_summary_data[emp.id]['missing_cp2'] +
                wa_summary_data[emp.id]['missing_pulang']
            )
        

    # DETERMINE ACTIVE TAB
    # Default to 'fingerprint'
    active_tab = 'fingerprint'
    
    # If requests explicit tab via GET or if WA Employee ID is present -> WhatsApp
    if request.GET.get('tab') == 'whatsapp' or wa_employee_id or request.GET.get('active_tab') == 'whatsapp':
        active_tab = 'whatsapp'

    # Convert keys to string for template compatibility
    print(f"DEBUG: Pre-Context. Matrix Size: {len(matrix_data)}. FP Stats Size: {len(fp_summary_stats)}")
    fp_summary_stats_str = {str(k): v for k, v in fp_summary_stats.items()} if fp_summary_stats else {}
    wa_summary_data_str = {str(k): v for k, v in wa_summary_data.items()} if wa_summary_data else {}

    context = {
        'selected_location': selected_location,
        'all_locations': all_locations,
        'can_select_location': can_select_location,
        'month_names': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'selected_month': month,
        'year_options': range(now.year - 2, now.year + 2),
        'selected_year': year,
        'fingerprint_employees': fingerprint_employees,
        'wa_employees': wa_employees,
        'matrix_data': matrix_data,
        'fp_summary_stats': fp_summary_stats_str,
        'wa_summary_data': wa_summary_data_str,
        'date_range': date_range,
        'holidays_map': holidays_map,
        'selected_employee_id': employee_id,
        'mode': mode,
        'personal_employee': personal_employee,
        'personal_logs': personal_logs,
        'summary_stats': summary_stats,
        # WA Context
        'wa_matrix_data': wa_matrix_data,
        'wa_summary_data': wa_summary_data,
        'wa_date_range': wa_date_range,
        'wa_categories': WA_CATEGORIES,
        'wa_mode': wa_mode,
        'wa_personal_employee': wa_personal_employee,
        'selected_wa_employee_id': wa_employee_id,
        # Active Tab State
        'active_tab': active_tab,
        # Monthly Report Workflow
        'report': report,
        'monthly_report_obj': report, # UNIQUE NAME
        'is_locked': is_locked,
        'error_message': error_message,
        'debug_report_str': str(report) if report else 'None',
    }
    
    print(f"[DEBUG] Final Context Report: {context.get('report')} (Type: {type(context.get('report'))})")
    print(f"[DEBUG] Final Context Error: {context.get('error_message')}")

    try:
        with open('d:\\dev\\attendance_core\\debug_trace.txt', 'a') as f:
            f.write(f"[{timezone.now()}] FINAL CONTEXT - Report: {context.get('report')} | Type: {type(context.get('report'))} | Error: {context.get('error_message')}\n")
    except:
        pass
        
    # WA Personal View (Monthly - Single Employee)
    if wa_employee_id:
        try:
            wa_target_employee = Employee.objects.get(id=wa_employee_id)
            wa_personal_employee = wa_target_employee  # Set for template context
            _, days_in_month = monthrange(year, month)
            wa_date_range = [date(year, month, day) for day in range(1, days_in_month + 1)]
                
            # Fetch all logs for this employee in the month (Remove strict source_type filter)
            wa_month_logs = AttendanceLog.objects.filter(
                employee=wa_target_employee,
                timestamp__year=year,
                timestamp__month=month
            ).order_by('timestamp')
            
            # Group by date and category
            wa_grouped = defaultdict(lambda: {cat: None for cat in WA_CATEGORIES})
            for log in wa_month_logs:
                local_dt = timezone.localtime(log.timestamp)
                log_date = local_dt.date()
                cat = log.log_category if log.log_category in WA_CATEGORIES else 'UNKNOWN'
                if cat in WA_CATEGORIES:
                    wa_grouped[log_date][cat] = {
                        'time': local_dt.strftime('%H:%M'),
                        'log': log,
                    }
            
            # Initialize Summary Stats for WA
            wa_summary_stats = {
                'late': 0, 'overtime': 0, 'alpha': 0, 'permit': 0, 'sick': 0,
                'missing_pagi': 0, 'missing_cp1': 0, 'missing_istirahat': 0,
                'missing_cp2': 0, 'missing_pulang': 0, 'total_missing': 0
            }
            
            # Schedule Constants (Reuse standard or define here)
            # Assuming standard 08:00 - 17:00 for calculation
            SCHEDULE_IN = timezone.datetime.strptime('08:00', '%H:%M').time()
            SCHEDULE_OUT = timezone.datetime.strptime('17:00', '%H:%M').time()
            dummy_date = date(2000, 1, 1)

            # Build personal logs list
            for d in wa_date_range:
                is_holiday = d in id_holidays
                is_weekend = d.weekday() == 6  # Only Sunday is weekend
                holiday_name = id_holidays.get(d)
                
                day_data = wa_grouped.get(d, {cat: None for cat in WA_CATEGORIES})
                
                # Count completed checkpoints
                completed = sum(1 for cat in WA_CATEGORIES if day_data.get(cat))
                progress = int((completed / 5) * 100)
                
                # Determine Status and Late/Overtime
                status = '-'
                status_class = ''
                late_minutes = 0
                
                # Check for explicit log status (SICK/PERMIT/ALPHA from AttendanceLog)
                # We check if any log in day_data has a special status
                special_status_log = None
                for cat in WA_CATEGORIES:
                    entry = day_data.get(cat)
                    if entry and entry.get('log'):
                        log = entry['log']
                        if log.status in ['SICK', 'PERMIT', 'ALPHA']:
                            special_status_log = log
                            break
                
                if special_status_log:
                    st = special_status_log.status
                    if st == 'SICK':
                        status = 'Sakit'
                        status_class = 'text-warning fw-bold'
                        wa_summary_stats['sick'] += 1
                    elif st == 'PERMIT':
                        status = 'Izin'
                        status_class = 'text-info fw-bold'
                        wa_summary_stats['permit'] += 1
                    elif st == 'ALPHA':
                        status = 'Alpha'
                        status_class = 'text-danger fw-bold'
                        wa_summary_stats['alpha'] += 1
                
                # Check for approved Leave (Sakit/Izin/Cuti) from EmployeeLeave or Alpha/Holiday/Weekend
                elif not day_data.get('MASUK'):
                    leave = EmployeeLeave.objects.filter(
                        employee=wa_personal_employee,
                        start_date__lte=d,
                        end_date__gte=d
                    ).first()
                    
                    if leave:
                        status = leave.get_leave_type_display()
                        if leave.leave_type == 'SAKIT':
                            status_class = 'text-warning fw-bold'
                            wa_summary_stats['sick'] += 1
                        elif leave.leave_type == 'IZIN':
                            status_class = 'text-info fw-bold'
                            wa_summary_stats['permit'] += 1
                        elif leave.leave_type == 'CUTI':
                            status_class = 'text-success fw-bold'
                        else:
                            status_class = 'text-secondary'
                    else:
                        # If NO leave found, check for Holiday, Weekend, or Alpha
                        if is_holiday:
                            status = 'Libur: ' + (holiday_name or '')
                            status_class = 'text-secondary'
                        elif is_weekend:
                            status = 'Minggu'
                            status_class = 'text-secondary'
                        elif d < now.date():
                            status = 'Alpha'
                            status_class = 'text-danger fw-bold'
                            wa_summary_stats['alpha'] += 1
                        else:
                            status = '-' # Future date

                elif day_data.get('MASUK'):
                    status = 'Hadir'
                    status_class = 'text-success fw-bold'
                    
                    # Calculate Late
                    try:
                        check_in_time = datetime.strptime(day_data['MASUK']['time'], '%H:%M').time()
                        if check_in_time > SCHEDULE_IN and not is_holiday and not is_weekend:
                            dt_in = timezone.datetime.combine(dummy_date, check_in_time)
                            dt_sch = timezone.datetime.combine(dummy_date, SCHEDULE_IN)
                            diff = (dt_in - dt_sch).total_seconds() / 60
                            late_minutes = int(diff)
                            wa_summary_stats['late'] += late_minutes
                    except (ValueError, TypeError):
                        pass
                        
                    # Calculate Overtime (if PULANG exists)
                    if day_data.get('PULANG'):
                        try:
                            check_out_time = datetime.strptime(day_data['PULANG']['time'], '%H:%M').time()
                            if check_out_time > SCHEDULE_OUT:
                                dt_out = timezone.datetime.combine(dummy_date, check_out_time)
                                dt_sch_out = timezone.datetime.combine(dummy_date, SCHEDULE_OUT)
                                diff = (dt_out - dt_sch_out).total_seconds() / 60
                                wa_summary_stats['overtime'] += int(diff)
                        except (ValueError, TypeError):
                            pass

                    if is_holiday:
                         status += ' (Hari Libur)'
                
                # Check for missing checkpoints (skip future, holidays, weekends)
                if not (d > now.date() or is_holiday or is_weekend):
                    if not day_data.get('MASUK'): wa_summary_stats['missing_pagi'] += 1
                    if not day_data.get('CHECKPOINT_1'): wa_summary_stats['missing_cp1'] += 1
                    if not day_data.get('ISTIRAHAT'): wa_summary_stats['missing_istirahat'] += 1
                    if not day_data.get('CHECKPOINT_2'): wa_summary_stats['missing_cp2'] += 1
                    if not day_data.get('PULANG'): wa_summary_stats['missing_pulang'] += 1

                wa_personal_logs.append({
                    'date': d,
                    'is_holiday': is_holiday,
                    'is_weekend': is_weekend,
                    'holiday_name': holiday_name,
                    'checkpoints': day_data,
                    'completed': completed,
                    'progress': progress,
                    'status': status,
                    'status_class': status_class,
                    'late_minutes': late_minutes,
                })
            
            # Calculate Total Missing
            wa_summary_stats['total_missing'] = (
                wa_summary_stats['missing_pagi'] +
                wa_summary_stats['missing_cp1'] +
                wa_summary_stats['missing_istirahat'] +
                wa_summary_stats['missing_cp2'] +
                wa_summary_stats['missing_pulang']
            )
        except Employee.DoesNotExist:
            wa_mode = 'daily'

    # Generate month/year options
    month_names = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]
    year_options = list(range(now.year - 2, now.year + 2))
    
    
    # DETERMINE ACTIVE TAB
    # Default to 'fingerprint'
    active_tab = 'fingerprint'
    
    # If requests explicit tab via GET or if WA Employee ID is present -> WhatsApp
    if request.GET.get('tab') == 'whatsapp' or wa_employee_id or request.GET.get('active_tab') == 'whatsapp':
        active_tab = 'whatsapp'
    
    context = {
        'page_title': 'Meja Kerja Rekapitulasi',
        'selected_location': selected_location,
        'all_locations': all_locations,
        'can_select_location': can_select_location,
        'user_role': user_role,
        'all_departments': all_departments,
        'selected_department_id': selected_department_id,
        'employees': employees,
        'matrix_data': matrix_data,
        'date_range': date_range,
        'selected_month': month,
        'selected_year': year,
        'month_names': month_names,
        'year_options': year_options,
        'holidays_map': holidays_map, 
        
        # Monthly Report Workflow
        'report': report,
        'is_locked': is_locked,
        'error_message': error_message,

        # Personal Mode Context (Fingerprint)
        'mode': mode,
        'personal_logs': personal_logs,
        'summary_stats': summary_stats,
        'selected_employee_id': str(employee_id) if employee_id else '',
        'personal_employee': personal_employee,  # Employee object for personal mode
        
        # Fingerprint Employees (for Tab 1)
        'fingerprint_employees': fingerprint_employees,
        'fp_summary_stats': fp_summary_stats_str, 
        'fp_count': fp_count if 'fp_count' in locals() else len(fingerprint_employees),
        'wa_count': wa_count if 'wa_count' in locals() else len(wa_employees),
        
        # WhatsApp / Telegram Context (for Tab 2)
        'wa_employees': wa_employees,
        'wa_mode': wa_mode,
        'wa_specific_date': wa_specific_date,
        'wa_daily_data': wa_daily_data,
        'wa_matrix_data': wa_matrix_data,
        'wa_date_range': wa_date_range if wa_employees else date_range,
        'wa_personal_logs': wa_personal_logs,
        'wa_categories': WA_CATEGORIES,
        'wa_category_icons': WA_CATEGORY_ICONS,
        'wa_summary_data': wa_summary_data if wa_employees else {},
        'wa_summary_stats': wa_summary_stats if 'wa_summary_stats' in locals() else {'late': 0, 'overtime': 0, 'alpha': 0, 'permit': 0, 'sick': 0},
        'selected_wa_employee_id': str(wa_employee_id) if wa_employee_id else '',
        'wa_personal_employee': wa_personal_employee,  # WA Employee object for personal mode
        'employee_search': employee_search,
        'wa_employee_search': wa_employee_search,
        
        # Role-based access
        'user_role': profile.role if profile else 'KERANI',
        
        # Active Tab State
        'active_tab': active_tab, 
        
        # Pagination Context
        'per_page': per_page if 'per_page' in locals() else 20,
        'page_obj': page_obj if 'page_obj' in locals() else None,
    }
    
    return render(request, 'portal/recap_matrix.html', context)


# =============================================================================
# EDIT ATTENDANCE MODAL (HTMX)
# =============================================================================

@login_required
def edit_attendance_modal(request, employee_id, date_str):
    """
    HTMX: Load modal untuk edit/create attendance log.
    
    Args:
        employee_id: UUID karyawan
        date_str: Format YYYY-MM-DD
    """
    from attendance.models import Employee, AttendanceLog
    from datetime import datetime
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('<div class="alert alert-danger">Format tanggal tidak valid</div>')
    
    # Check if log exists for this employee and date
    existing_log = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__date=target_date
    ).first()
    
    # Status options
    status_choices = [
        ('CHECKIN', 'Hadir (H)'),
        ('SICK', 'Sakit (S)'),
        ('PERMIT', 'Izin (I)'),
        ('ALPHA', 'Alpha (A)'),
    ]
    
    context = {
        'employee': employee,
        'target_date': target_date,
        'date_str': date_str,
        'existing_log': existing_log,
        'status_choices': status_choices,
        'is_edit': existing_log is not None,
    }
    
    return render(request, 'portal/partials/_modal_edit_log.html', context)


@login_required
def save_attendance_cell(request):
    """
    HTMX POST: Simpan/Update attendance log dan return cell fragment.
    
    Returns only the updated cell HTML for hx-swap="outerHTML".
    """
    from attendance.models import Employee, AttendanceLog
    from datetime import datetime
    from django.utils import timezone
    
    if request.method != 'POST':
        return HttpResponse('<div class="alert alert-warning">Method not allowed</div>')
    
    employee_id = request.POST.get('employee_id')
    date_str = request.POST.get('date_str')
    status = request.POST.get('status')
    clock_in = request.POST.get('clock_in')
    clock_out = request.POST.get('clock_out')
    notes = request.POST.get('notes', '')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (Employee.DoesNotExist, ValueError) as e:
        return HttpResponse(f'<div class="status-cell cell-empty">!</div>')
        
    # =========================================================================
    # LOCKING CHECK
    # =========================================================================
    from attendance.models import MonthlyReport
    is_locked = False
    try:
        # Check if report exists and is locked
        report = MonthlyReport.objects.filter(
            location=employee.home_base, # Assumes employee assigned to this location
            period_month=target_date.month,
            period_year=target_date.year
        ).first()
        
        if report and report.status != 'DRAFT' and report.status != 'REQ_REVISI':
            is_locked = True
            return HttpResponse('<div class="status-cell cell-empty" title="Laporan Terkunci"><i class="fas fa-lock"></i></div>')
            
    except Exception:
        pass # Fail safe, allow edit if check fails (or handle stricter)

    # Determine display status code
    status_map = {
        'CHECKIN': 'H',
        'SICK': 'S',
        'PERMIT': 'I',
        'ALPHA': 'A',
    }
    display_status = status_map.get(status, 'A')
    
    # Check if log exists
    existing_log = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__date=target_date
    ).first()
    
    now = timezone.now()
    
    if status == 'ALPHA':
        # Delete existing log if marking as Alpha
        if existing_log:
            existing_log.delete()
    else:
        # Create or update log
        if existing_log:
            existing_log.status = status
            existing_log.notes = notes
            existing_log.save()
        else:
            # Create new log
            # Parse clock_in time or use default 08:00
            if clock_in:
                try:
                    time_parts = clock_in.split(':')
                    log_datetime = timezone.make_aware(
                        datetime.combine(target_date, datetime.strptime(clock_in, '%H:%M').time())
                    )
                except:
                    log_datetime = timezone.make_aware(
                        datetime.combine(target_date, datetime.strptime('08:00', '%H:%M').time())
                    )
            else:
                log_datetime = timezone.make_aware(
                    datetime.combine(target_date, datetime.strptime('08:00', '%H:%M').time())
                )
            
            AttendanceLog.objects.create(
                employee=employee,
                timestamp=log_datetime,
                status=status,
                source_type='MANUAL',
                notes=notes,
                captured_at=now,
            )
    
    # Return cell fragment with updated status
    context = {
        'employee': employee,
        'target_date': target_date,
        'date_str': date_str,
        'status': display_status,
        'is_holiday': target_date.weekday() >= 5, # Simple check, ideally use holiday lib
        'is_locked': is_locked,
    }
    
    return render(request, 'portal/partials/_attendance_cell.html', context)


# =============================================================================
# WA CELL EDIT (HTMX) - Role-Based Access Control
# =============================================================================

@login_required
def wa_edit_cell(request, employee_id, date_str, category):
    """
    HTMX: Load modal content untuk edit WA attendance cell.
    Admin bisa set waktu, Kerani hanya bisa set Izin/Sakit.
    """
    from attendance.models import Employee, AttendanceLog
    from datetime import datetime
    
    employee = Employee.objects.get(id=employee_id)
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Get user role
    profile = getattr(request.user, 'employee_profile', None)
    user_role = profile.role if profile else 'KERANI'
    
    # Check if log exists for this cell
    existing_log = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__date=target_date,
        log_category=category,
        source_type='TELEGRAM'
    ).first()
    
    existing_time = existing_log.timestamp.strftime('%H:%M') if existing_log else ''
    existing_notes = existing_log.notes if existing_log else ''
    
    # Category labels
    category_labels = {
        'MASUK': 'Pagi',
        'CHECKPOINT_1': 'Checkpoint 1',
        'ISTIRAHAT': 'Istirahat',
        'CHECKPOINT_2': 'Checkpoint 2',
        'PULANG': 'Pulang',
    }
    
    context = {
        'employee': employee,
        'date_str': date_str,
        'category': category,
        'category_label': category_labels.get(category, category),
        'user_role': user_role,
        'existing_time': existing_time,
        'existing_notes': existing_notes,
        'is_admin': user_role in ['ADMIN', 'HRD'],
        'cell_id': f"wa-cell-{employee.id}-{date_str}-{category}", # Target ID for HTMX
    }
    
    return render(request, 'portal/partials/_wa_edit_cell_modal.html', context)


@login_required
@require_POST
@transaction.atomic
def wa_save_cell(request):
    """
    HTMX: Save WA attendance cell.
    Admin bisa set waktu, Kerani hanya bisa set Izin/Sakit.
    Updates EmployeeLeave for Izin/Sakit to ensure sync with Personal View.
    """
    from attendance.models import Employee, AttendanceLog, EmployeeLeave
    from django.utils import timezone
    from datetime import datetime
    from attendance.models import MonthlyReport # Added import for MonthlyReport
    
    employee_id = request.POST.get('employee_id')
    date_str = request.POST.get('date_str')
    category = request.POST.get('category')
    action = request.POST.get('action')  # 'set_time', 'izin', 'sakit', 'delete'
    time_value = request.POST.get('time_value', '')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (Employee.DoesNotExist, ValueError):
        return HttpResponse('<div class="status-cell cell-empty">!</div>')

    # =========================================================================
    # LOCKING CHECK
    # =========================================================================
    try:
        # Check if report exists and is locked
        report = MonthlyReport.objects.filter(
            location=employee.home_base,
            period_month=target_date.month,
            period_year=target_date.year
        ).first()
        
        if report and report.status != 'DRAFT' and report.status != 'REQ_REVISI':
             # Return Locked Icon (Context-aware: inside a td, but here we return cell content usually)
             # Wait, wa_save_cell returns... let's check what it returns.
             # It likely returns the cell fragment.
             return HttpResponse('<span class="text-muted"><i class="fas fa-lock"></i></span>')
            
    except Exception:
        pass 
    
    # Get user role
    profile = getattr(request.user, 'employee_profile', None)
    user_role = profile.role if profile else 'KERANI'
    
    # 1. Handle DELETE action
    if action == 'delete':
        # Delete specific log for this category
        AttendanceLog.objects.filter(
            employee=employee,
            timestamp__date=target_date,
            log_category=category,
            source_type='TELEGRAM'
        ).delete()
        
        # Also check if we should delete leave if it exists and we are clicking delete on a cell
        # (Optional: might depend on business logic, for now focused on log)
        
        return HttpResponse('<span class="text-muted opacity-25">-</span>')
    
    # 2. Handle IZIN/SAKIT (Create EmployeeLeave)
    if action in ['izin', 'sakit']:
        # Clear existing logs for this day to avoid ambiguity (optional, but cleaner)
        # AttendanceLog.objects.filter(employee=employee, timestamp__date=target_date).delete()
        
        # Create/Update EmployeeLeave
        leave_type_map = {'izin': 'IZIN', 'sakit': 'SAKIT'}
        l_type = leave_type_map.get(action, 'IZIN')
        
        # Check if leave already exists for this day
        existing_leave = EmployeeLeave.objects.filter(
            employee=employee,
            start_date__lte=target_date,
            end_date__gte=target_date
        ).first()
        
        if existing_leave:
            existing_leave.leave_type = l_type
            existing_leave.save()
        else:
            EmployeeLeave.objects.create(
                employee=employee,
                leave_type=l_type,
                start_date=target_date,
                end_date=target_date,
                notes=f'Manual Update via WA Matrix ({category})',
                created_by=request.user
            )
            
        return HttpResponse(f'<span class="text-warning fw-bold" style="font-size: 0.6rem;">{action.upper()[:2]}</span>')

    # 3. Handle SET TIME (Admin Only)
    elif action == 'set_time':
        if user_role not in ['ADMIN', 'HRD']:
            return HttpResponse('<span class="text-danger">Tidak diizinkan</span>')
            
        log_time = time_value if time_value else '08:00'
        notes = 'Manual Entry'
        
        # Calculate timestamp
        try:
            log_datetime = timezone.make_aware(
                datetime.combine(target_date, datetime.strptime(log_time, '%H:%M').time())
            )
        except:
            log_datetime = timezone.make_aware(
                datetime.combine(target_date, datetime.strptime('08:00', '%H:%M').time())
            )
            
        # Update or Create Log
        existing_log = AttendanceLog.objects.filter(
            employee=employee,
            timestamp__date=target_date,
            log_category=category,
            source_type='TELEGRAM'
        ).first()
        
        if existing_log:
            existing_log.timestamp = log_datetime
            existing_log.notes = notes
            existing_log.save()
        else:
            AttendanceLog.objects.create(
                employee=employee,
                timestamp=log_datetime,
                log_category=category,
                source_type='TELEGRAM',
                notes=notes,
            )
            
        cell_html = f'<span class="text-success" style="font-size: 0.6rem;">{log_time}</span>'
    
    # Validation fallback
    if 'cell_html' not in locals():
        if action in ['izin', 'sakit']:
            cell_html = f'<span class="text-warning fw-bold" style="font-size: 0.6rem;">{action.upper()[:2]}</span>'
        else:
            cell_html = '<span class="text-danger">Error</span>'
            
    # 3. RECALCULATE MONTHLY STATS (Reactive Update - HTMX OOB)
    # We need to import the stats function from views (it's defined in this file)
    # Since it is a function in the same module, we can just call it if it's in scope, 
    # but strictly speaking Python function ordering matters if not hoisted. 
    # django views are usually all loaded.
    
    # Re-calculate stats
    stats = get_employee_monthly_stats(employee, target_date.year, target_date.month)
    
    oob_html = ""
    for key, val in stats.items():
        td_id = f"stats-{employee.id}-{key}"
        
        # Match template styles
        style = ""
        cls = "text-center fw-bold bg-light"
        
        if key == 'H': 
            style = 'border-left: 2px solid #dee2e6; color: #198754;'
        elif key == 'T': 
            style = 'color: #fd7e14;'
        elif key == 'A': 
            style = 'color: #dc3545;'
        elif key == 'S': 
            cls += " text-warning"
        elif key == 'I': 
            cls += " text-info"
            
        oob_html += f'<td id="{td_id}" hx-swap-oob="true" class="{cls}" style="{style}">{val}</td>'
        
    # Add script to close modal
    close_script = '<script>bootstrap.Modal.getInstance(document.getElementById("waEditModal")).hide();</script>'
        
    return HttpResponse(cell_html + oob_html + close_script)


# =============================================================================
# EMPLOYEE LEAVE CRUD ENDPOINTS
# =============================================================================

@login_required
@require_POST
def employee_leave_add(request):
    """Add a new leave record for an employee"""
    from attendance.models import Employee, EmployeeLeave
    from django.utils import timezone
    from datetime import datetime
    
    employee_id = request.POST.get('employee_id')
    leave_type = request.POST.get('leave_type')
    start_date_str = request.POST.get('start_date')
    end_date_str = request.POST.get('end_date')
    notes = request.POST.get('notes', '')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        if end_date < start_date:
            return HttpResponse('<div class="alert alert-danger">Tanggal selesai harus setelah tanggal mulai</div>')
            
        # =====================================================================
        # LOCKING CHECK
        # =====================================================================
        from attendance.models import MonthlyReport
        lock_report = MonthlyReport.objects.filter(
            location=employee.home_base,
            period_month=start_date.month,
            period_year=start_date.year
        ).first()
        if lock_report and lock_report.status not in ('DRAFT', 'REQ_REVISI'):
            return HttpResponse('<div class="alert alert-danger">Laporan bulan ini sudah dikunci. Tidak bisa menambah data.</div>')
            
        leave = EmployeeLeave.objects.create(
            employee=employee,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            notes=notes,
            created_by=request.user
        )
        
        if 'attachment' in request.FILES:
            leave.attachment = request.FILES['attachment']
            leave.save()
        
        # Determine target container from request headers (HTMX)
        target_id = request.headers.get('HX-Target', 'leave-list-container')

        # Return updated leave list HTML
        return render(request, 'portal/_employee_leave_list.html', {
            'leaves': employee.leaves.all()[:10],
            'employee': employee,
            'target_id': target_id
        })
        
    except Employee.DoesNotExist:
        return HttpResponse('<div class="alert alert-danger">Karyawan tidak ditemukan</div>')
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Error: {str(e)}</div>')


@login_required
@require_POST
def employee_leave_delete(request, leave_id):
    """Delete a leave record"""
    from attendance.models import EmployeeLeave
    
    try:
        leave = EmployeeLeave.objects.get(id=leave_id)
        employee = leave.employee
        
        # =====================================================================
        # LOCKING CHECK
        # =====================================================================
        from attendance.models import MonthlyReport
        lock_report = MonthlyReport.objects.filter(
            location=employee.home_base,
            period_month=leave.start_date.month,
            period_year=leave.start_date.year
        ).first()
        if lock_report and lock_report.status not in ('DRAFT', 'REQ_REVISI'):
            return HttpResponse('<div class="alert alert-danger">Laporan bulan ini sudah dikunci. Tidak bisa menghapus data.</div>')
            
        leave.delete()
        
        # Determine target container from request headers (HTMX)
        target_id = request.headers.get('HX-Target', 'leave-list-container')

        # Return updated leave list HTML
        return render(request, 'portal/_employee_leave_list.html', {
            'leaves': employee.leaves.all()[:10],
            'employee': employee,
            'target_id': target_id
        })

        
    except EmployeeLeave.DoesNotExist:
        return HttpResponse('<div class="alert alert-danger">Data tidak ditemukan</div>')


@login_required
def employee_leave_edit(request, leave_id):
    """Load modal content for editing a leave record"""
    from attendance.models import EmployeeLeave
    
    try:
        leave = EmployeeLeave.objects.get(id=leave_id)
        return render(request, 'portal/partials/_leave_edit_modal.html', {'leave': leave})
    except EmployeeLeave.DoesNotExist:
        return HttpResponse('<div class="alert alert-danger">Data tidak ditemukan</div>')


@login_required
@require_POST
def employee_leave_update(request, leave_id):
    """Update a leave record"""
    from attendance.models import EmployeeLeave
    from datetime import datetime
    
    try:
        leave = EmployeeLeave.objects.get(id=leave_id)
        employee = leave.employee
        
        # Update fields
        leave.leave_type = request.POST.get('leave_type')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        leave.notes = request.POST.get('notes', '')
        
        leave.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        leave.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # =====================================================================
        # LOCKING CHECK
        # =====================================================================
        from attendance.models import MonthlyReport
        # Check Original Date
        lock_report_old = MonthlyReport.objects.filter(
            location=employee.home_base,
            period_month=leave.start_date.month,
            period_year=leave.start_date.year
        ).first()
        if lock_report_old and lock_report_old.status not in ('DRAFT', 'REQ_REVISI'):
             return HttpResponse('<div class="alert alert-danger">Laporan bulan asal sudah dikunci. Tidak bisa mengubah data.</div>')
             
        # Check New Date (if changing months)
        if leave.start_date.month != leave.start_date.month or leave.start_date.year != leave.start_date.year:
            lock_report_new = MonthlyReport.objects.filter(
                location=employee.home_base,
                period_month=leave.start_date.month,
                period_year=leave.start_date.year
            ).first()
            if lock_report_new and lock_report_new.status not in ('DRAFT', 'REQ_REVISI'):
                 return HttpResponse('<div class="alert alert-danger">Laporan bulan tujuan sudah dikunci. Tidak bisa memindahkan data.</div>')

        if leave.end_date < leave.start_date:
            return HttpResponse('<div class="alert alert-danger">Tanggal selesai harus setelah tanggal mulai</div>')
            
        # Handle file upload if provided
        if 'attachment' in request.FILES:
            leave.attachment = request.FILES['attachment']
            
        leave.save()
        
        # Determine target container from request headers (HTMX)
        target_id = request.headers.get('HX-Target', 'leave-list-container')

        # Return updated leave list HTML
        return render(request, 'portal/_employee_leave_list.html', {
            'leaves': employee.leaves.all()[:10],
            'employee': employee,
            'target_id': target_id
        })
        
    except EmployeeLeave.DoesNotExist:
        return HttpResponse('<div class="alert alert-danger">Data tidak ditemukan</div>')
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Error: {str(e)}</div>')


@login_required
def employee_leave_list(request, employee_id):
    """Get leave list for an employee (HTMX)"""
    from attendance.models import Employee
    
    try:
        employee = Employee.objects.get(id=employee_id)
        # Determine target container from request headers (HTMX)
        target_id = request.headers.get('HX-Target', 'leave-list-container')

        return render(request, 'portal/_employee_leave_list.html', {
            'leaves': employee.leaves.all()[:10],
            'employee': employee,
            'target_id': target_id
        })
    except Employee.DoesNotExist:
        return HttpResponse('<div class="alert alert-warning">Karyawan tidak ditemukan</div>')
def get_employee_monthly_stats(employee, year, month):
    """
    Calculate monthly stats for a single employee (H, T, A, S, I).
    Used for partial updates (HTMX OOB).
    """
    from attendance.models import AttendanceLog, EmployeeLeave
    from calendar import monthrange
    from datetime import date, timedelta, datetime
    import holidays
    from django.utils import timezone
    try:
        from attendance.utils import get_employee_schedule
    except ImportError:
        get_employee_schedule = None

    stats = {'H': 0, 'A': 0, 'T': 0, 'S': 0, 'I': 0}
    
    _, days_in_month = monthrange(year, month)
    date_range = [date(year, month, day) for day in range(1, days_in_month + 1)]
    id_holidays = holidays.ID(years=[year, year-1])
    
    # 1. Fetch Logs
    logs = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__year=year,
        timestamp__month=month
    ).order_by('timestamp')
    
    date_logs = {}
    for log in logs:
        local_dt = timezone.localtime(log.timestamp)
        d = local_dt.date()
        if d not in date_logs: date_logs[d] = []
        date_logs[d].append(log)
        
    # 2. Fetch Leaves
    leaves = EmployeeLeave.objects.filter(
        employee=employee,
        start_date__lte=date_range[-1],
        end_date__gte=date_range[0]
    )
    leave_map = {}
    for l in leaves:
        curr = max(l.start_date, date_range[0])
        end = min(l.end_date, date_range[-1])
        while curr <= end:
            leave_map[curr] = l.leave_type
            curr += timedelta(days=1)
            
    now = timezone.now().date()
    DEFAULT_SCHEDULE_IN = datetime.strptime('08:00', '%H:%M').time()
    
    for d in date_range:
        day_logs_list = date_logs.get(d, [])
        is_holiday = d in id_holidays
        is_weekend = d.weekday() == 6
        
        if day_logs_list:
            first_log = day_logs_list[0]
            status_map = {
                'CHECKIN': 'H', 'HADIR': 'H',
                'SICK': 'S', 'SAKIT': 'S',
                'PERMIT': 'I', 'IZIN': 'I',
                'ALPHA': 'A'
            }
            st = status_map.get(first_log.status, 'H')
            
            if st == 'H':
                stats['H'] += 1
                # Late logic
                if get_employee_schedule:
                    daily_sch, _ = get_employee_schedule(employee, d)
                    sch_in = daily_sch.clock_in if daily_sch else DEFAULT_SCHEDULE_IN
                    tol = daily_sch.late_tolerance if daily_sch else 0
                    
                    # Dummy date logic
                    dummy_date = date(2000, 1, 1)
                    dt_sch = datetime.combine(dummy_date, sch_in)
                    dt_threshold = dt_sch + timedelta(minutes=tol)
                    
                    local_dt = timezone.localtime(first_log.timestamp)
                    clock_time = local_dt.time().replace(second=0, microsecond=0)
                    threshold_time = dt_threshold.time().replace(second=0, microsecond=0)
                    
                    if clock_time > threshold_time and not is_holiday and not is_weekend:
                        stats['T'] += 1
            elif st == 'S': stats['S'] += 1
            elif st == 'I': stats['I'] += 1
            elif st == 'A': stats['A'] += 1
            
        else:
            # Check leave
            l_type = leave_map.get(d)
            if l_type:
                code_map = {'IZIN': 'I', 'SAKIT': 'S', 'CUTI': 'C', 'CUTI_KHUSUS': 'C'}
                st = code_map.get(l_type, 'I')
                if st == 'C': st = 'I'
                if st in stats: stats[st] += 1
            elif is_holiday: pass
            elif is_weekend: pass
            elif d < now:
                stats['A'] += 1
                
    return stats

@login_required
def fp_edit_cell(request, employee_id, date_str):
    """
    HTMX: Load modal content untuk edit Fingerprint attendance cell.
    Allows setting Hadir, Sakit, Izin, Alpha.
    """
    from attendance.models import Employee, AttendanceLog, EmployeeLeave
    from datetime import datetime
    
    employee = Employee.objects.get(id=employee_id)
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Check for existing log (Checkin/Checkout)
    existing_log = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__date=target_date,
        source_type='FINGERPRINT'
    ).first()
    
    # Check for existing leave
    existing_leave = EmployeeLeave.objects.filter(
        employee=employee,
        start_date__lte=target_date,
        end_date__gte=target_date
    ).first()
    
    existing_status = ''
    if existing_leave:
        existing_status = existing_leave.get_leave_type_display()
    elif existing_log:
        existing_status = 'Hadir (Data Mesin)'
        if existing_log.notes == 'Manual Entry':
            existing_status = 'Hadir (Manual)'
            
    context = {
        'employee': employee,
        'date_str': date_str,
        'existing_status': existing_status,
        'existing_log': existing_log,
    }
    
    return render(request, 'portal/partials/_fp_edit_cell_modal.html', context)


@login_required
@require_POST
@transaction.atomic
def fp_save_cell(request):
    """
    HTMX: Save Fingerprint attendance cell.
    Updates AttendanceLog (for Hadir) or EmployeeLeave (for Izin/Sakit).
    """
    from attendance.models import Employee, AttendanceLog, EmployeeLeave
    from django.utils import timezone
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    
    employee_id = request.POST.get('employee_id')
    date_str = request.POST.get('date_str')
    action = request.POST.get('action')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # =====================================================================
        # LOCKING CHECK - Prevent edits on locked reports
        # =====================================================================
        from attendance.models import MonthlyReport
        lock_report = MonthlyReport.objects.filter(
            location=employee.home_base,
            period_month=target_date.month,
            period_year=target_date.year
        ).first()
        if lock_report and lock_report.status not in ('DRAFT', 'REQ_REVISI'):
            return HttpResponse(
                '<div class="status-cell cell-empty" title="Laporan Terkunci">'
                '<i class="fas fa-lock"></i></div>'
            )
        
        # 1. CLEAR EXISTING STATUS (Delete relevant logs/leaves for this day)
        # Delete Manual Logs
        AttendanceLog.objects.filter(
            employee=employee,
            timestamp__date=target_date,
            source_type='FINGERPRINT',
            notes='Manual Entry'
        ).delete()
        
        # Check for Leave covering this date
        # Note: We delete any leave intersecting this day for simplicity in this cell-based edit mode
        EmployeeLeave.objects.filter(
            employee=employee,
            start_date__lte=target_date,
            end_date__gte=target_date
        ).delete() 
    
        # 2. APPLY NEW ACTION
        cell_content = "+"
        css_class = "cell-empty"
        
        if action == 'delete':
             cell_content = "+"
             css_class = "cell-empty"
            
        elif action == 'hadir':
            # Create Dummy Log (Checkin 08:00, Checkout 17:00)
            log_in = timezone.make_aware(datetime.combine(target_date, datetime.strptime('08:00', '%H:%M').time()))
            log_out = timezone.make_aware(datetime.combine(target_date, datetime.strptime('17:00', '%H:%M').time()))
            
            AttendanceLog.objects.create(
                employee=employee, timestamp=log_in, source_type='FINGERPRINT', device_id='MANUAL', notes='Manual Entry'
            )
            AttendanceLog.objects.create(
                employee=employee, timestamp=log_out, source_type='FINGERPRINT', device_id='MANUAL', notes='Manual Entry'
            )
            cell_content = "H"
            css_class = "cell-hadir"
            
        elif action in ['izin', 'sakit', 'cuti']:
            leave_type_map = {'izin': 'IZIN', 'sakit': 'SAKIT', 'cuti': 'CUTI'}
            l_type = leave_type_map.get(action, 'IZIN')
            EmployeeLeave.objects.create(
                employee=employee,
                leave_type=l_type,
                start_date=target_date,
                end_date=target_date,
                notes='Manual Update via Matrix'
            )
            
            # Map back to status code for CSS
            status_code_map = {'sakit': 'S', 'izin': 'I', 'cuti': 'C'}
            code = status_code_map.get(action, 'I')
            cell_content = code
            css_class = "cell-izin"
            
        elif action == 'alpha':
            # Ensure no logs/leaves exist (already cleared above)
            cell_content = "A"
            css_class = "cell-alpha"
            
        cell_html = f'<div class="status-cell {css_class}">{cell_content}</div>'
        
        # 3. RECALCULATE MONTHLY STATS (Reactive Update - HTMX OOB)
        stats = get_employee_monthly_stats(employee, target_date.year, target_date.month)
        
        oob_html = ""
        for key, val in stats.items():
            td_id = f"stats-{employee.id}-{key}"
            
            # Match template styles
            style = ""
            cls = "text-center fw-bold bg-light"
            
            if key == 'H': 
                style = 'border-left: 2px solid #dee2e6; color: #198754;'
            elif key == 'T': 
                style = 'color: #fd7e14;'
            elif key == 'A': 
                style = 'color: #dc3545;'
            elif key == 'S': 
                cls += " text-warning"
            elif key == 'I': 
                cls += " text-info"
                
            oob_html += f'<td id="{td_id}" hx-swap-oob="true" class="{cls}" style="{style}">{val}</td>'
            
        return HttpResponse(cell_html + oob_html)
        
    except Exception as e:
        return HttpResponse(f'<div class="status-cell cell-empty" title="Error: {str(e)}">!</div>')


# =============================================================================
# PUBLISH REPORT (Monthly Report Workflow)
# =============================================================================

@login_required
def publish_report(request, report_id):
    if request.method == "POST":
        from attendance.models import MonthlyReport, ReportHistory
        from django.shortcuts import get_object_or_404
        from django.utils import timezone
        
        report = get_object_or_404(MonthlyReport, id=report_id)
        
        # Validasi sederhana (Opsional: Cek permission user disini)
        
        # Update Status
        report.status = 'SUBMITTED'
        report.submitted_by = request.user
        report.submitted_at = timezone.now()
        report.save()
        
        # Catat History
        ReportHistory.objects.create(
            report=report,
            actor=request.user,
            action="SUBMITTED",
            new_status="SUBMITTED",
            comment="Laporan dipublish oleh Kerani (Final)"
        )
        
        # CRITICAL FIX: Return response kosong dengan header trigger refresh
        return HttpResponse(status=200, headers={'HX-Refresh': 'true'})
    
    return HttpResponse(status=400)


# =============================================================================
# VERIFY REPORT (HRD Workflow)
# =============================================================================

@login_required
def verify_report(request, report_id):
    if request.method == "POST":
        from attendance.models import MonthlyReport, ReportHistory
        from django.shortcuts import get_object_or_404
        from django.utils import timezone
        
        report = get_object_or_404(MonthlyReport, id=report_id)
        
        # Check Permissions: Only HRD or SUPERVISOR or Superuser
        profile = getattr(request.user, 'employee_profile', None)
        is_hrd = profile and profile.role in ['HRD', 'SUPERVISOR']
        
        if not (request.user.is_superuser or is_hrd):
             return HttpResponse("Unauthorized", status=403)

        # SUPERVISOR LOCATION CHECK
        if profile and profile.role == 'SUPERVISOR':
             # Ensure report location matches assigned location (or is descendant)
             # Simplify: Just check if report.location is in user's allowed locations
             if profile.assigned_location:
                 allowed_locations = profile.assigned_location.get_descendants(include_self=True)
                 if report.location not in allowed_locations:
                     return HttpResponse("Unauthorized Location", status=403)
             else:
                 # Supervisor without location cannot verify anything
                 return HttpResponse("Unauthorized: No Assigned Location", status=403)

        # Update Status
        report.status = 'VERIFIED'
        report.verified_by = request.user
        report.verified_at = timezone.now()
        report.save()
        
        # History
        ReportHistory.objects.create(
            report=report,
            actor=request.user,
            action="VERIFIED",
            new_status="VERIFIED",
            comment="Laporan diverifikasi oleh HRD"
        )
        
        # Return success with refresh header (HTMX compatible)
        return HttpResponse(status=200, headers={'HX-Refresh': 'true'})
    
    return HttpResponse(status=400)


# =============================================================================
# UNLOCK REQUEST WORKFLOW (Kerani <-> HRD)
# =============================================================================

@login_required
def request_unlock(request, report_id):
    """
    Kerani initiates a request to unlock a SUBMITTED report.
    Status changes: SUBMITTED -> REQ_UNLOCK
    """
    if request.method == "POST":
        
        report = get_object_or_404(MonthlyReport, id=report_id)

        
        # Validation: Only SUBMITTED can be requested to unlock
        if report.status != 'SUBMITTED':
             return HttpResponse("Invalid status for unlock request", status=400)

        # Update Status
        report.status = 'REQ_UNLOCK'
        report.save()
        
        # History
        ReportHistory.objects.create(
            report=report,
            actor=request.user,
            action="REQUEST_UNLOCK",
            new_status="REQ_UNLOCK",
            comment="Permintaan buka kunci oleh Kerani"
        )
        
        return HttpResponse(status=200, headers={'HX-Refresh': 'true'})
    return HttpResponse(status=400)

@login_required
def approve_unlock(request, report_id):
    """
    HRD approves unlock request.
    Status changes: REQ_UNLOCK -> REQ_REVISI (Editable)
    """
    if request.method == "POST":
        from attendance.models import MonthlyReport, ReportHistory
        from django.shortcuts import get_object_or_404
        
        # Permission Check
        profile = getattr(request.user, 'employee_profile', None)
        is_hrd = profile and profile.role == 'HRD'
        if not (request.user.is_superuser or is_hrd):
             return HttpResponse("Unauthorized", status=403)

        report = get_object_or_404(MonthlyReport, id=report_id)
        
        if report.status != 'REQ_UNLOCK':
             return HttpResponse("Invalid status for approval", status=400)

        # Update Status -> REQ_REVISI (Allows editing)
        report.status = 'REQ_REVISI'
        report.save()
        
        # History
        ReportHistory.objects.create(
            report=report,
            actor=request.user,
            action="APPROVE_UNLOCK",
            new_status="REQ_REVISI",
            comment="Permintaan buka kunci disetujui HRD"
        )
        
        return HttpResponse(status=200, headers={'HX-Refresh': 'true'})
    return HttpResponse(status=400)

@login_required
def reject_unlock(request, report_id):
    """
    HRD rejects unlock request.
    Status changes: REQ_UNLOCK -> SUBMITTED (Locked again)
    """
    if request.method == "POST":
        from attendance.models import MonthlyReport, ReportHistory
        from django.shortcuts import get_object_or_404
        
        # Permission Check
        is_hrd = request.user.is_superuser or (
            hasattr(request.user, 'employee_profile') and 
            request.user.employee_profile.role == 'HRD'
        )
        if not is_hrd:
             return HttpResponse("Unauthorized", status=403)

        report = get_object_or_404(MonthlyReport, id=report_id)
        
        if report.status != 'REQ_UNLOCK':
             return HttpResponse("Invalid status for rejection", status=400)

        # Update Status -> SUBMITTED (Back to locked state)
        report.status = 'SUBMITTED'
        report.save()
        
        # History
        ReportHistory.objects.create(
            report=report,
            actor=request.user,
            action="REJECT_UNLOCK",
            new_status="SUBMITTED",
            comment="Permintaan buka kunci ditolak HRD"
        )
        
        return HttpResponse(status=200, headers={'HX-Refresh': 'true'})
    return HttpResponse(status=400)


# =============================================================================
# EXPORT EXCEL (REPORT)
# =============================================================================

@login_required
def export_matrix_excel(request, report_id):
    """
    Export Laporan Bulanan ke Excel (Format Accounting).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from attendance.models import MonthlyReport, Employee, AttendanceLog, EmployeeLeave
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from django.utils import timezone
    from calendar import monthrange
    from datetime import date, timedelta, datetime, time
    import holidays

    # 1. Fetch Report & Meta
    report = get_object_or_404(MonthlyReport, id=report_id)
    location = report.location
    year = report.period_year
    month = report.period_month
    
    # 2. Setup Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fingerprint"
    
    # Styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFeb9c", end_color="FFeb9c", fill_type="solid") # Light Yellow
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 3. Build Header Row
    # Col A-C Fixed
    headers = ["NIK", "Nama Karyawan", "Jabatan"]
    
    # Col D... Days
    _, days_in_month = monthrange(year, month)
    date_cols = []
    for d in range(1, days_in_month + 1):
        date_cols.append(str(d))
        headers.append(str(d))
        
    # Col Summary
    summary_headers = ["H", "T", "A", "S", "I", "Terlambat (m)"]
    headers.extend(summary_headers)
    
    ws.append(headers)
    
    # Apply Style to Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        
    # Set Column Widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    
    # Date Cols Width
    for i, _ in enumerate(date_cols):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 4
        
    # 4. Fetch Data (Optimized)
    employees = Employee.objects.filter(home_base=location, is_active=True).order_by('full_name')
    
    logs = AttendanceLog.objects.filter(
        employee__home_base=location,
        timestamp__year=year,
        timestamp__month=month
    ).select_related('employee').order_by('timestamp')
    
    # Map Logs: emp_id -> day -> list of logs
    log_map = {}
    for log in logs:
        eid = log.employee_id
        day = timezone.localtime(log.timestamp).day
        if eid not in log_map: log_map[eid] = {}
        if day not in log_map[eid]: log_map[eid][day] = []
        log_map[eid][day].append(log)
        
    # Map Leaves
    leaves = EmployeeLeave.objects.filter(
        employee__home_base=location,
        start_date__lte=date(year, month, days_in_month),
        end_date__gte=date(year, month, 1)
    )
    leave_map = {} # emp_id -> day -> type
    for l in leaves:
        eid = l.employee_id
        if eid not in leave_map: leave_map[eid] = {}
        curr = max(l.start_date, date(year, month, 1))
        end = min(l.end_date, date(year, month, days_in_month))
        while curr <= end:
            leave_map[eid][curr.day] = l.leave_type
            curr += timedelta(days=1)
            
    # Holidays
    id_holidays = holidays.ID(years=[year])
    
    # Import for schedule
    from attendance.utils import get_employee_schedule
    
    # =============================================================================
    # SHEET 1: FINGERPRINT (EXISTING LOGIC)
    # =============================================================================
    
    # 5. Write Data Rows (Fingerprint)
    for emp in employees.filter(attendance_method='FINGERPRINT'):
        row = [emp.nik or '-', emp.full_name, emp.get_employee_type_display() or '-']
        
        stats = {'H': 0, 'T': 0, 'A': 0, 'S': 0, 'I': 0, 'LateM': 0}
        
        for d in range(1, days_in_month + 1):
            day_date = date(year, month, d)
            is_holiday = day_date in id_holidays
            is_weekend = day_date.weekday() == 6
            
            cell_val = "-"
            
            emp_logs_day = log_map.get(emp.id, {}).get(d, [])
            emp_leave_day = leave_map.get(emp.id, {}).get(d)
            
            if emp_logs_day:
                # Logic Hadir/Telat
                # Prioritize 'MASUK' or first log (Check-In)
                # Logs are sorted by timestamp asc, so [0] is Check-In
                first_log = next((l for l in emp_logs_day if l.log_category == 'MASUK'), emp_logs_day[0])
                
                status_map = {'CHECKIN': 'H', 'HADIR': 'H', 'SICK': 'S', 'SAKIT': 'S', 'PERMIT': 'I', 'IZIN': 'I', 'ALPHA': 'A'}
                st = status_map.get(first_log.status, 'H')
                
                if st == 'H':
                    stats['H'] += 1
                    cell_val = "H"
                    
                    # Late Calculation (Dynamic)
                    try:
                        dt_log = timezone.localtime(first_log.timestamp)
                        log_time = dt_log.time()
                        
                        # Get Dynamic Schedule
                        daily_sch, _ = get_employee_schedule(emp, day_date)
                        
                        # Default Values
                        sch_in = time(8, 0)
                        tol = 5
                        
                        if daily_sch:
                            sch_in = daily_sch.clock_in
                            tol = daily_sch.late_tolerance
                            
                        # Threshold
                        dummy_date = date(2000, 1, 1)
                        dt_sch = datetime.combine(dummy_date, sch_in)
                        dt_threshold = dt_sch + timedelta(minutes=tol)
                        
                        # Truncate Seconds to match Web Matrix (Lenient)
                        log_time_trunc = log_time.replace(second=0, microsecond=0)
                        threshold_trunc = dt_threshold.time().replace(second=0, microsecond=0)
                        
                        # Exclude Holidays and Sundays from Late Count
                        if log_time_trunc > threshold_trunc and not is_holiday and not is_weekend:
                            stats['T'] += 1
                            # Calculate Diff
                            dt_in = datetime.combine(dummy_date, log_time)
                            diff = (dt_in - dt_threshold).total_seconds() / 60
                            stats['LateM'] += int(diff)
                            
                    except Exception as e:
                        print(f"Excel Calc Error: {e}")
                        pass

                    
                elif st == 'S': 
                     stats['S'] += 1
                     cell_val = "S"
                elif st == 'I': 
                     stats['I'] += 1
                     cell_val = "I"
                elif st == 'A': 
                     stats['A'] += 1
                     cell_val = "A"
                     
            elif emp_leave_day:
                 code_map = {'IZIN': 'I', 'SAKIT': 'S', 'CUTI': 'C', 'CUTI_KHUSUS': 'C'}
                 st = code_map.get(emp_leave_day, 'I')
                 if st == 'S': 
                     stats['S'] += 1
                     cell_val = "S"
                 elif st == 'I' or st == 'C': 
                     stats['I'] += 1
                     cell_val = "I" if st == 'I' else "C"
            elif is_holiday:
                cell_val = "L"
            elif is_weekend:
                cell_val = "M" 
            elif day_date < timezone.now().date():
                stats['A'] += 1
                cell_val = "A"
            else:
                 pass # Default "-"
            
            row.append(cell_val)
            
        # Add Summaries
        row.append(stats['H'])
        row.append(stats['T'])
        row.append(stats['A'])
        row.append(stats['S'])
        row.append(stats['I'])
        row.append(stats['LateM'])
        
        ws.append(row)
        
        # Style Data Row
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            if cell.column > 3:
                 cell.alignment = center_align

    # =============================================================================
    # SHEET 2: WHATSAPP LAPANGAN (NEW LOGIC)
    # =============================================================================
    
    ws2 = wb.create_sheet(title="WhatsApp Lapangan")
    
    # 1. Header Logic
    # Row 1: Fixed Headers + Summary Header + Date Headers
    # Col 1-3: NIK, Nama, Jabatan
    # Col 4-9: Σ Tidak Absen (Merged)
    # Col 10+: Dates (Merged 5)
    
    # Fixed Headers
    headers_wa = ['NIK', 'Nama Karyawan', 'Jabatan']
    for col_num, header in enumerate(headers_wa, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws2.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)

    # Summary Header (Σ Tidak Absen)
    summary_start_col = 4
    ws2.cell(row=1, column=summary_start_col, value="Σ Tidak Absen").font = header_font
    ws2.cell(row=1, column=summary_start_col).fill = header_fill # Check if fill works on merged cell top-left
    ws2.cell(row=1, column=summary_start_col).alignment = center_align
    ws2.cell(row=1, column=summary_start_col).border = thin_border
    ws2.merge_cells(start_row=1, start_column=summary_start_col, end_row=1, end_column=summary_start_col + 5)
    
    summary_subs = ['P', '1', 'I', '2', 'O', 'Iz']
    for i, sub in enumerate(summary_subs):
        c = ws2.cell(row=2, column=summary_start_col + i, value=sub)
        c.font = header_font
        c.alignment = center_align
        c.border = thin_border
        c.fill = PatternFill(start_color="FFeb9c", end_color="FFeb9c", fill_type="solid") # Light Yellow
        ws2.column_dimensions[get_column_letter(summary_start_col + i)].width = 4

    # Date Headers
    current_col = 10
    
    for day in range(1, days_in_month + 1):
        # Row 1: Tanggal (Merged 5 cells)
        date_cell = ws2.cell(row=1, column=current_col, value=f"{day}")
        date_cell.font = header_font
        date_cell.alignment = center_align
        date_cell.fill = header_fill
        date_cell.border = thin_border
        
        ws2.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 4)
        
        # Row 2: Sub-Headers
        sub_headers = ['M', 'C1', 'Ist', 'C2', 'P']
        for i, sub in enumerate(sub_headers):
            cell_idx = current_col + i
            sub_cell = ws2.cell(row=2, column=cell_idx, value=sub)
            sub_cell.alignment = center_align
            sub_cell.border = thin_border
            sub_cell.fill = header_fill 
            ws2.column_dimensions[get_column_letter(cell_idx)].width = 5
            
        current_col += 5

    # Data Rows WA
    wa_employees = employees.filter(attendance_method__in=['WHATSAPP', 'MANUAL'])
    
    wa_row_num = 3
    
    for emp in wa_employees:
        # Calculate Missing Stats & Row Data
        missing_stats = {
            'missing_pagi': 0, 
            'missing_cp1': 0, 
            'missing_istirahat': 0, 
            'missing_cp2': 0, 
            'missing_pulang': 0,
            'izin_count': 0
        }
        
        row_time_cells = [] # Stores (value, style) tuples or just values
        
        for d in range(1, days_in_month + 1):
            day_date = date(year, month, d)
            is_holiday = day_date in id_holidays
            is_weekend = day_date.weekday() == 6
            is_future = day_date > timezone.now().date()
            
            emp_logs_day = log_map.get(emp.id, {}).get(d, [])
            emp_leave_day = leave_map.get(emp.id, {}).get(d)
            
            # Init Times
            times = {'MASUK': '-', 'CHECKPOINT_1': '-', 'ISTIRAHAT': '-', 'CHECKPOINT_2': '-', 'PULANG': '-'}
            
            # Populate Times from Logs
            for l in emp_logs_day:
                cat = l.log_category
                t_str = timezone.localtime(l.timestamp).strftime('%H:%M')
                if cat in times:
                    times[cat] = t_str
            
            # Calculate Summary (Logic similar to view)
            if not is_future and not is_holiday and not is_weekend:
                 if emp_leave_day:
                     # Check leave logic (IZIN counts to izin_count, SAKIT also usually)
                     # For 'missing' counts, if on leave, we don't count as missing?
                     # Let's assume if leave exists, no missing penalty.
                     # Mapping based on codes:
                     if emp_leave_day in ['IZIN', 'SAKIT', 'CUTI', 'CUTI_KHUSUS']:
                         missing_stats['izin_count'] += 1
                 else:
                     # Check each checkpoint
                     if times['MASUK'] == '-': missing_stats['missing_pagi'] += 1
                     if times['CHECKPOINT_1'] == '-': missing_stats['missing_cp1'] += 1
                     if times['ISTIRAHAT'] == '-': missing_stats['missing_istirahat'] += 1
                     if times['CHECKPOINT_2'] == '-': missing_stats['missing_cp2'] += 1
                     if times['PULANG'] == '-': missing_stats['missing_pulang'] += 1
            
            # Append to row data
            ordered_keys = ['MASUK', 'CHECKPOINT_1', 'ISTIRAHAT', 'CHECKPOINT_2', 'PULANG']
            for k in ordered_keys:
                row_time_cells.append(times[k])

        # Write Row
        # 1. Info
        ws2.cell(row=wa_row_num, column=1, value=emp.nik or '-').border = thin_border
        ws2.cell(row=wa_row_num, column=2, value=emp.full_name).border = thin_border
        ws2.cell(row=wa_row_num, column=3, value=emp.get_employee_type_display() or '-').border = thin_border
        
        # 2. Stats
        ws2.cell(row=wa_row_num, column=4, value=missing_stats['missing_pagi']).border = thin_border
        ws2.cell(row=wa_row_num, column=5, value=missing_stats['missing_cp1']).border = thin_border
        ws2.cell(row=wa_row_num, column=6, value=missing_stats['missing_istirahat']).border = thin_border
        ws2.cell(row=wa_row_num, column=7, value=missing_stats['missing_cp2']).border = thin_border
        ws2.cell(row=wa_row_num, column=8, value=missing_stats['missing_pulang']).border = thin_border
        ws2.cell(row=wa_row_num, column=9, value=missing_stats['izin_count']).border = thin_border
        
        # Color stats cells if > 0
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for c_idx in range(4, 9): # P to O
             c = ws2.cell(row=wa_row_num, column=c_idx)
             c.alignment = center_align
             if isinstance(c.value, int) and c.value > 0:
                 c.fill = red_fill
        
        # Iz cell blue
        c_iz = ws2.cell(row=wa_row_num, column=9)
        c_iz.alignment = center_align
        if isinstance(c_iz.value, int) and c_iz.value > 0:
             c_iz.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid") # Light Blue

        # 3. Time Data
        current_data_col = 10
        for val in row_time_cells:
            c = ws2.cell(row=wa_row_num, column=current_data_col, value=val)
            c.alignment = center_align
            c.border = thin_border
            current_data_col += 1
            
        wa_row_num += 1

    # 6. Response
    filename = f"Laporan Absensi {location.name} Periode {month}-{year}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# =============================================================================
# PDF EXPORT
# =============================================================================

import qrcode
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import landscape, portrait, A4, LEGAL
from reportlab.lib.units import inch, cm, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

def generate_qr(data):
    """Generate in-memory QR code image"""
    qr = qrcode.QRCode(box_size=2, border=0)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer

@login_required
def export_report_pdf(request, report_id):
    report = get_object_or_404(MonthlyReport, id=report_id)
    location = report.location
    month = report.period_month
    year = report.period_year
    
    report_type = request.GET.get('type', 'main')
    buffer = BytesIO()

    # =========================================================================
    # TYPE: MAIN REPORT (Matrix Fingerprint) - LEGAL LANDSCAPE
    # =========================================================================
    if report_type == 'main':
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Fetch Employees (Fingerprint only or Mixed?)
        # Main report usually includes everyone
        employees = Employee.objects.filter(
            home_base=location,
            is_active=True
        ).order_by('full_name')

        total_employees = employees.count()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(LEGAL), # Changed to Legal
            rightMargin=0.5*cm,
            leftMargin=0.5*cm, # Wider margins for Legal
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # --- CORPORATE HEADER FUNCTION ---
        import os
        from django.conf import settings
        from datetime import time
        from django.utils import timezone
        
        def draw_corporate_header(canvas, doc):
            canvas.saveState()
            
            # Constants
            COMPANY_NAME = "PT. PETALING MANDRAGUNA"
            COMPANY_ADDRESS = "Jl. Hayam Wuruk, Kota Jambi, Jambi"
            COMPANY_CONTACT = "Telp: (0741) 33264 | Email: hr@mandraguna.com"
            
            # Colors
            corpo_blue = colors.Color(31/255, 58/255, 147/255)
            
            # Dimensions
            page_width, page_height = doc.pagesize
            # Landscape Legal: 35.56 cm width, 21.59 cm height
            
            # 1. Logo
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'company_logo.png')
            
            # Position: Top Left
            # Page Height is ~21.6cm. Top Margin is 1cm.
            # We want header in top 3.5cm.
            
            logo_x = 1.0*cm
            logo_y = page_height - 3.5*cm
            logo_w = 2.5*cm
            logo_h = 2.5*cm
            
            try:
                canvas.drawImage(logo_path, logo_x, logo_y, width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)
                text_x = logo_x + logo_w + 0.5*cm
            except Exception:
                text_x = 1.0*cm # Fallback if no logo
            
            # 2. Text
            # 2. Text
            # Name
            canvas.setFont("Helvetica-Bold", 16)
            canvas.setFillColor(corpo_blue)
            canvas.drawCentredString(page_width / 2.0, logo_y + 1.8*cm, COMPANY_NAME)
            
            # Address
            canvas.setFont("Helvetica", 10)
            canvas.setFillColor(colors.darkgrey)
            canvas.drawCentredString(page_width / 2.0, logo_y + 1.3*cm, COMPANY_ADDRESS)
            
            # Contact
            canvas.drawCentredString(page_width / 2.0, logo_y + 0.9*cm, COMPANY_CONTACT)
            
            # 3. Separator
            canvas.setLineWidth(1.5)
            canvas.setStrokeColor(corpo_blue)
            sep_y = logo_y - 0.2*cm
            canvas.line(1.0*cm, sep_y, page_width - 1.0*cm, sep_y)
            
            canvas.restoreState()
            
        elements = []
        
        # SPACER to push content down (Corporate Header takes approx 3.5cm)
        elements.append(Spacer(1, 3.5*cm))
        
        styles = getSampleStyleSheet()
        style_center = ParagraphStyle(name='Center', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=14)
        style_normal = styles['Normal']
        
        # Signatures Preparation
        sub_time = report.submitted_at.strftime("%d-%m-%Y %H:%M") if report.submitted_at else ""
        ver_time = report.verified_at.strftime("%d-%m-%Y %H:%M") if report.verified_at else ""
        app_time = report.approved_at.strftime("%d-%m-%Y %H:%M") if report.approved_at else ""
        
        qr_sub_img = Image(generate_qr(f"VALID:SUBMITTED:{report.id}"), width=25*mm, height=25*mm) if report.submitted_by else Paragraph("", style_normal)
        qr_ver_img = Image(generate_qr(f"VALID:VERIFIED:{report.id}"), width=25*mm, height=25*mm) if report.verified_by else Paragraph("", style_normal)
        qr_app_img = Image(generate_qr(f"VALID:APPROVED:{report.id}"), width=25*mm, height=25*mm) if report.approved_by else Paragraph("", style_normal)
        
        sig_data = [
            ["Dibuat Oleh,", "Diverifikasi Oleh,", "Disetujui Oleh,"],
            [qr_sub_img, qr_ver_img, qr_app_img],
            [
                report.submitted_by.username if report.submitted_by else "(....................)",
                report.verified_by.username if report.verified_by else "(....................)",
                report.approved_by.username if report.approved_by else "(....................)"
            ],
            [
                f"Tgl: {sub_time}" if sub_time else "",
                f"Tgl: {ver_time}" if ver_time else "",
                f"Tgl: {app_time}" if app_time else ""
            ],
            ["Kerani", "HRD / Personalia", "Accounting/Manager"]
        ]
        # Title
        elements.append(Paragraph(f"BERITA ACARA REKAPITULASI ABSENSI", style_center))
        elements.append(Paragraph(f"{location.name.upper()}", style_center))
        elements.append(Spacer(1, 0.5*cm))
        
        # Periode Info
        elements.append(Paragraph(f"Periode: {calendar.month_name[month]} {year}", style_center))
        elements.append(Spacer(1, 0.5*cm))
        
        # Summary Table
        start_date = date(year, month, 1)
        end_date = date(year, month, days_in_month)
        
        ids = employees.values_list('id', flat=True)
        logs = AttendanceLog.objects.filter(
            employee_id__in=ids,
            timestamp__date__range=[start_date, end_date]
        )
        present_count = logs.values('employee', 'timestamp__date').distinct().count()
        fp_count = logs.filter(source_type='FINGERPRINT').values('employee', 'timestamp__date').distinct().count()
        
        # Prepare Leave Map for H.T.A.S.I logic
        leaves = EmployeeLeave.objects.filter(
            employee__in=employees,
            start_date__lte=end_date,
            end_date__gte=start_date
        )
        from collections import defaultdict
        leave_map = defaultdict(lambda: defaultdict(str))
        for l in leaves:
            curr = max(l.start_date, start_date)
            end = min(l.end_date, end_date)
            while curr <= end:
                leave_map[l.employee_id][curr] = l.leave_type
                curr += timedelta(days=1)
                
        # Calculate WA Count (exclude FINGERPRINT)
        wa_count = logs.exclude(source_type='FINGERPRINT').values('employee', 'timestamp__date').distinct().count()

        summary_data = [
            ['Total Karyawan', f"{total_employees}"],
            ['Total Man-Days (Kehadiran)', f"{present_count}"],
            ['Kehadiran via Fingerprint', f"{fp_count}"],
            ['Kehadiran via WA/Manual', f"{wa_count}"],
            ['Total Revisi', f"{report.total_unlock_requests}x"],
            ['Status Laporan', f"{report.get_status_display().upper()}"],
        ]
        
        
        t_summary = Table(summary_data, colWidths=[6*cm, 6*cm], hAlign='CENTER')
        t_summary.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 2*cm))
        
        # Signatures
        # Row 1: Titles
        t_sig_cover = Table(sig_data, colWidths=[6*cm, 6*cm, 6*cm])
        t_sig_cover.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        elements.append(t_sig_cover)
        elements.append(PageBreak())
        
        # --- PAGE 2 ONWARDS: MATRIX ---
        # Add Spacer for Header on subsequent pages too?
        # If we use onFirstPage only, then subsequent pages won't have it.
        # But `doc.build` has `onLaterPages`.
        # Usually Matrix page doesn't need big header, but "BERITA ACARA" title implies cover page.
        # Let's keep matrix page clean or add simplified header?
        # User only asked for "header korporat sebelum judul utama laporan".
        # So mainly first page.
        
        # However, if we put `onFirstPage`, we need `onLaterPages` if we want it there too.
        # Ideally, matrix page has minimal header title.
        
        # --- PAGE 2 ONWARDS: MATRIX ---
        
        # Split Employees by Method
        employees_finger = employees.filter(attendance_method='FINGERPRINT')
        employees_wa = employees.exclude(attendance_method='FINGERPRINT') # Catch-all for WA/Manual
        
        groups = [
            ("SUMBER: MESIN FINGERPRINT", employees_finger),
            ("SUMBER: WHATSAPP / TELEGRAM / MANUAL", employees_wa)
        ]
        
        # Helper for holidays
        id_holidays = holidays.Indonesia(years=year)
        
        for title, emp_group in groups:
            if not emp_group.exists():
                continue
                
            # Title Source
            elements.append(Paragraph(title, style_center))
            elements.append(Spacer(1, 0.2*cm))
            
            header_row = ['No', 'Nama'] + [str(d) for d in range(1, days_in_month + 1)] + ['H', 'T', 'A', 'S', 'I']
            data = [header_row]
            
            for idx, emp in enumerate(emp_group, 1):
                row = [str(idx), emp.full_name]
                emp_logs = logs.filter(employee=emp)
                
                stats = {'H': 0, 'T': 0, 'A': 0, 'S': 0, 'I': 0}

                for d in range(1, days_in_month + 1):
                    day_date = date(year, month, d)
                    is_holiday = day_date in id_holidays
                    is_weekend = day_date.weekday() == 6
                    # FIX: Use order_by('timestamp') to get the FIRST log (Check-In)
                    # Default ordering is -timestamp (Check-Out), which caused False Late.
                    day_log = emp_logs.filter(timestamp__date=day_date).order_by('timestamp').first()
                    val = '-'
                    
                    # Check Leave First (Priority)
                    leave_type = leave_map.get(emp.id, {}).get(day_date)
                    if leave_type:
                        if leave_type == 'SAKIT': val = 'S'; stats['S'] += 1
                        elif leave_type in ['IZIN', 'CUTI', 'CUTI_KHUSUS']: val = 'I'; stats['I'] += 1
                        else: val = 'I'; stats['I'] += 1 # Fallback
                    
                    elif day_log:
                        # Log Check
                        if day_log.status == 'SAKIT':
                            val = 'S'; stats['S'] += 1
                        elif day_log.status == 'IZIN':
                            val = 'I'; stats['I'] += 1
                        else:
                            # Dynamic Late Check
                            # Import helper locally to avoid circular imports or ensure availability
                            from attendance.utils import get_employee_schedule
                            from datetime import datetime, timedelta
                            
                            val = 'H'
                            stats['H'] += 1
                            
                            log_dt = timezone.localtime(day_log.timestamp)
                            log_time = log_dt.time()
                            
                            # Get Schedule
                            daily_sch, _ = get_employee_schedule(emp, day_date)
                            
                            # Std defaults
                            sch_in = time(8, 0)
                            tol = 5
                            
                            if daily_sch:
                                sch_in = daily_sch.clock_in
                                tol = daily_sch.late_tolerance
                            
                            # Calc Threshold
                            dummy_date = date(2000, 1, 1)
                            dt_sch = datetime.combine(dummy_date, sch_in)
                            dt_threshold = dt_sch + timedelta(minutes=tol)
                            
                            # Truncate Seconds to match Web Matrix (Lenient)
                            log_time_trunc = log_time.replace(second=0, microsecond=0)
                            threshold_trunc = dt_threshold.time().replace(second=0, microsecond=0)
                            
                            # Exclude Holidays and Sundays from Late Count
                            if log_time_trunc > threshold_trunc and not is_holiday and not is_weekend:
                                # Late!
                                stats['T'] += 1
                                # val remains 'H' to match Excel/Web matrix
                                # If user WANTS 'T' in cell, we can change this back.
                                # But standardizing means 'H' in cell, 'T' in summary.
                                # User said "matrix finger... tidak sinkron", and matrix uses 'H' usually.
                                pass
                    else:
                        # No Log, No Leave
                        if day_date.weekday() == 6: val = 'L' # Sunday
                        elif day_date in id_holidays: val = 'L' # Holiday
                        else: 
                            val = 'A'; stats['A'] += 1

                    row.append(val)
                
                # Append Stats Columns
                row.extend([str(stats['H']), str(stats['T']), str(stats['A']), str(stats['S']), str(stats['I'])])
                
                data.append(row)
            
            # Column Widths for Legal Landscape (~33cm usable)
            col_widths = [1*cm, 5*cm] + [0.7*cm] * days_in_month + [0.8*cm] * 5
            
            t_matrix = Table(data, repeatRows=1, colWidths=col_widths)
            t_matrix.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (2,0), (-1,-1), 'CENTER'), # Center data days
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), # Header bg
                ('LEFTPADDING', (0,0), (-1,-1), 2),
                ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ]))
            elements.append(t_matrix)
            elements.append(Spacer(1, 1*cm)) # Space between tables
        
        filename = f"Laporan_Utama_{location.name}_{month}-{year}.pdf"

        # Build with Header
        doc.build(elements, onFirstPage=draw_corporate_header)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # =========================================================================
    # TYPE: WA REPORT (Summary Table) - A4 PORTRAIT
    # =========================================================================
    elif report_type == 'wa':
        days_in_month = calendar.monthrange(year, month)[1]
        start_date = date(year, month, 1)
        end_date = date(year, month, days_in_month)

        doc = SimpleDocTemplate(
            buffer,
            pagesize=portrait(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(name='Title', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=14)
        style_sub = ParagraphStyle(name='Sub', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=10)
        style_normal = styles['Normal']
        
        # 1. Header
        elements.append(Paragraph(f"REKAPITULASI ABSENSI LAPANGAN (WHATSAPP)", style_title))
        elements.append(Paragraph(f"Lokasi: {location.name} | Periode: {calendar.month_name[month]} {year}", style_sub))
        elements.append(Spacer(1, 0.5*cm))
        
        # 2. Fetch Data (Optimized)
        from django.db.models import Q
        # Use robust filter for WA employees (same as dashboard)
        wa_employees = Employee.objects.filter(
            home_base=location,
            is_active=True
        ).order_by('full_name')

        # Filter actually relevant employees
        # Check logs directly for WA/Telegram indicators, regardless of employee's current setting.
        has_wa_logs = set(AttendanceLog.objects.filter(
            timestamp__year=year,
            timestamp__month=month,
            employee__home_base=location
        ).filter(
            Q(source_type='TELEGRAM') | Q(verification_method='WA') | Q(employee__attendance_method__in=['WHATSAPP', 'WA'])
        ).values_list('employee_id', flat=True))

        # Filter employees list to only those with WA logs OR method=WA
        relevant_employees = [
            e for e in wa_employees 
            if e.id in has_wa_logs or e.attendance_method in ['WHATSAPP', 'WA']
        ]
        
        # Fetch Logs Bulk
        wa_logs = AttendanceLog.objects.filter(
            timestamp__date__range=[start_date, end_date],
            employee__in=relevant_employees
        ).select_related('employee')
        
        # Map Logs: EmpID -> Date -> Cat -> Time
        from collections import defaultdict
        data_map = defaultdict(lambda: defaultdict(dict))
        for log in wa_logs:
            d = log.timestamp.date()
            cat = log.log_category
            time_str = timezone.localtime(log.timestamp).strftime('%H:%M')
            data_map[log.employee_id][d][cat] = time_str
            # Also capture status/notes if needed? 
            # Dashboard logic uses log.status for H/S/I/A
            if not data_map[log.employee_id][d].get('STATUS'):
                data_map[log.employee_id][d]['STATUS'] = log.status

        # Fetch Leaves Bulk
        leaves = EmployeeLeave.objects.filter(
            employee__in=relevant_employees,
            start_date__lte=end_date,
            end_date__gte=start_date
        )
        leave_map = defaultdict(lambda: defaultdict(str))
        for l in leaves:
            curr = max(l.start_date, start_date)
            end = min(l.end_date, end_date)
            while curr <= end:
                leave_map[l.employee_id][curr] = l.leave_type
                curr += timedelta(days=1)
        
        # Holidays
        id_holidays = holidays.ID(years=[year, year-1])
        
        # 3. Build Summary Data
        summary_rows = []
        # Header: No, Nama, M, C1, Ist, C2, P, H, T, A, S, I
        table_header = ['No', 'Nama Karyawan', 'M', 'C1', 'Ist', 'C2', 'P', 'H', 'T', 'A', 'S', 'I']
        summary_rows.append(table_header)
        
        WA_CATS = ['MASUK', 'CHECKPOINT_1', 'ISTIRAHAT', 'CHECKPOINT_2', 'PULANG']
        
        for idx, emp in enumerate(relevant_employees, 1):
            stats = {
                'M': 0, 'C1': 0, 'Ist': 0, 'C2': 0, 'P': 0,
                'H': 0, 'T': 0, 'A': 0, 'S': 0, 'I': 0
            }
            
            # Iterate dates
            curr = start_date
            while curr <= end_date:
                is_holiday = curr in id_holidays or curr.weekday() == 6
                day_data = data_map.get(emp.id, {}).get(curr, {})
                leave_type = leave_map.get(emp.id, {}).get(curr)
                
                # --- STATUS COUNTS (H, T, A, S, I) ---
                status_code = None
                
                # Priority 1: Leaves
                if leave_type:
                    if leave_type in ['SAKIT']: 
                        stats['S'] += 1; status_code = 'S'
                    elif leave_type in ['IZIN', 'CUTI', 'CUTI_KHUSUS']: 
                        stats['I'] += 1; status_code = 'I'
                
                # Priority 2: Logs
                elif day_data.get('MASUK'):
                    stats['H'] += 1; status_code = 'H'
                    # Late Check
                    if day_data['MASUK'] > "08:00":
                        stats['T'] += 1
                
                # Priority 3: Alpha (If past date, not holiday)
                elif curr < timezone.now().date() and not is_holiday:
                    stats['A'] += 1; status_code = 'A'
                
                # --- MISSED CHECKPOINTS ---
                # Only count if NOT Holiday/Weekend AND NOT Izin/Sakit
                if not is_holiday and status_code not in ['S', 'I', 'A']: # If Alpha, maybe count as missed? Dashboard counts missed for Alpha too usually.
                    # Let's follow dashboard logic: count if missing and expected
                    # Dashboard: if not day_data.get('MASUK'): wa_summary_data['missing_pagi'] += 1
                    # But dashboard logic runs on ALL days except future/holidays
                    
                    if not day_data.get('MASUK'): stats['M'] += 1
                    if not day_data.get('CHECKPOINT_1'): stats['C1'] += 1
                    if not day_data.get('ISTIRAHAT'): stats['Ist'] += 1
                    if not day_data.get('CHECKPOINT_2'): stats['C2'] += 1
                    if not day_data.get('PULANG'): stats['P'] += 1

                curr += timedelta(days=1)
            
            row = [
                str(idx),
                emp.full_name[:25], # Truncate name
                str(stats['M']), str(stats['C1']), str(stats['Ist']), str(stats['C2']), str(stats['P']),
                str(stats['H']), str(stats['T']), str(stats['A']), str(stats['S']), str(stats['I'])
            ]
            summary_rows.append(row)
            
        # 4. Create Table
        # Col Widths: Total A4 width ~ 19cm available (21 - margin)
        # No: 1cm, Name: 6cm, Others: 1cm each (10 cols = 10cm) -> Total 17cm. Fits.
        cw_summary = [1*cm, 6*cm] + [1*cm]*10
        t_summary = Table(summary_rows, colWidths=cw_summary, repeatRows=1)
        
        t_summary.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'), # All center except name?
            ('ALIGN', (1,0), (1,-1), 'LEFT'),   # Name Left
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), # Header Bold
        ]))
        
        elements.append(t_summary)
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Spacer(1, 0.5*cm))

        # Legend Box (Requested Style)
        # Header + 2 Rows of 5 Columns
        legend_data = [
            ['KETERANGAN KODE:', '', '', '', ''], # Header merged
            ['M : Masuk (Pagi)', 'C1 : Checkpoint 1', 'Ist : Istirahat', 'H : Hadir', 'S : Sakit'],
            ['P : Pulang (Sore)', 'C2 : Checkpoint 2', 'T : Terlambat', 'A : Alpha', 'I : Izin']
        ]
        
        # Calculate proportional width (Total A4 width ~19cm - margins)
        # A4 = 21cm, Margins L/R = 1cm each -> 19cm usable
        # 19cm / 5 cols = 3.8 cm per col
        col_w = (19*cm) / 5
        t_legend = Table(legend_data, colWidths=[col_w]*5)
        
        legend_style = TableStyle([
            # General Style
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Oblique'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.darkgrey),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            
            # Header Style
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('SPAN', (0,0), (-1,0)), # Merge first row
            ('ALIGN', (0,0), (-1,0), 'LEFT'),
            
            # Box & Background
            ('BACKGROUND', (0,0), (-1,-1), HexColor('#f9f9f9')),
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey), # Outer border
            ('GRID', (0,0), (-1,-1), 0, colors.white), # No inner grid (or white/invisible)
            
            # Padding
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ])
        t_legend.setStyle(legend_style)
        elements.append(t_legend)
        
        elements.append(Spacer(1, 1*cm))
        
        elements.append(Spacer(1, 1*cm))
        
        # 5. Signatures (Copy from Main Report logic)
        # Verify timestamps
        sub_time = report.submitted_at.strftime("%d-%m-%Y %H:%M") if report.submitted_at else ""
        ver_time = report.verified_at.strftime("%d-%m-%Y %H:%M") if report.verified_at else ""
        app_time = report.approved_at.strftime("%d-%m-%Y %H:%M") if report.approved_at else ""
        
        qr_sub_img = Image(generate_qr(f"VALID:SUBMITTED:{report.id}"), width=25*mm, height=25*mm) if report.submitted_by else Paragraph("", style_normal)
        qr_ver_img = Image(generate_qr(f"VALID:VERIFIED:{report.id}"), width=25*mm, height=25*mm) if report.verified_by else Paragraph("", style_normal)
        qr_app_img = Image(generate_qr(f"VALID:APPROVED:{report.id}"), width=25*mm, height=25*mm) if report.approved_by else Paragraph("", style_normal)
        
        sig_data = [
            ["Dibuat Oleh,", "Diverifikasi Oleh,", "Disetujui Oleh,"],
            [qr_sub_img, qr_ver_img, qr_app_img],
            [
                report.submitted_by.username if report.submitted_by else "(....................)",
                report.verified_by.username if report.verified_by else "(....................)",
                report.approved_by.username if report.approved_by else "(....................)"
            ],
            [
                f"Tgl: {sub_time}" if sub_time else "",
                f"Tgl: {ver_time}" if ver_time else "",
                f"Tgl: {app_time}" if app_time else ""
            ],
            ["Kerani", "HRD / Personalia", "Accounting/Manager"]
        ]
        
        t_sig = Table(sig_data, colWidths=[6*cm, 6*cm, 6*cm])
        t_sig.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        
        elements.append(t_sig)
        
        filename = f"Laporan_WA_{location.name}_{month}-{year}.pdf"

    else:
        return HttpResponse("Invalid Report Type")

    # 6. Generator Info
    gen_info = f"Dicetak oleh: {request.user.username} pada {timezone.now().strftime('%d-%m-%Y %H:%M:%S')}"
    elements.append(Paragraph(gen_info, ParagraphStyle(name='Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# =============================================================================
# ACCOUNTING ACTIONS
# =============================================================================

@login_required
def approve_payment(request, report_id):
    """
    Accounting approve payment -> Finalize Report.
    """
    # 1. Permission Check
    profile = getattr(request.user, 'employee_profile', None)
    is_accounting = profile and profile.role == 'ACCOUNTING'
    is_supervisor = profile and profile.role == 'SUPERVISOR'
    
    if not (request.user.is_superuser or is_accounting or is_supervisor):
        from django.contrib import messages
        messages.error(request, "Akses ditolak. Hanya Accounting atau Supervisor yang dapat menyetujui pembayaran.")
        return redirect('portal:dashboard')

    # 2. Get Report
    from attendance.models import MonthlyReport, ReportHistory
    from django.utils import timezone
    report = get_object_or_404(MonthlyReport, id=report_id)
    
    # SUPERVISOR LOCATION CHECK
    if is_supervisor and not request.user.is_superuser:
        if profile.assigned_location:
             allowed_locations = profile.assigned_location.get_descendants(include_self=True)
             if report.location not in allowed_locations:
                 from django.contrib import messages
                 messages.error(request, "Akses ditolak. Lokasi laporan tidak sesuai dengan otoritas Anda.")
                 return redirect('portal:dashboard')
        else:
             from django.contrib import messages
             messages.error(request, "Akun Supervisor Anda tidak memiliki lokasi tugas.")
             return redirect('portal:dashboard')

    # 3. Validate Status
    if report.status != 'VERIFIED':
        from django.contrib import messages
        messages.error(request, f"Hanya laporan berstatus VERIFIED yang dapat disetujui. Status saat ini: {report.get_status_display()}")
        return redirect('portal:dashboard')

    # 4. Update Status & Stamps
    old_status = report.status
    report.status = 'APPROVED'
    report.approved_by = request.user
    report.approved_at = timezone.now()
    report.save()

    # 5. Log History
    ReportHistory.objects.create(
        report=report,
        actor=request.user,
        action='APPROVE_PAYMENT',
        previous_status=old_status,
        new_status='APPROVED',
        comment="Pembayaran disetujui oleh Accounting (Final)."
    )

    from django.contrib import messages
    messages.success(request, f"Pembayaran untuk {report} telah disetujui. Laporan kini berstatus FINAL/LUNAS.")
    return redirect('portal:dashboard')


# =============================================================================
# PERSONAL PRINT FEATURE
# =============================================================================

@login_required
def print_employee_modal(request, employee_id):
    """
    HTMX: Load modal for selecting Month/Year for Personal Report.
    """
    from attendance.models import Employee
    from django.utils import timezone
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Default to current month
    now = timezone.now()
    years = range(now.year, now.year - 2, -1) # Last 2 years
    
    context = {
        'emp': employee,
        'current_month': now.month,
        'current_year': now.year,
        'years': years,
        'months': range(1, 13)
    }
    return render(request, 'portal/partials/_print_employee_modal.html', context)

@login_required
def export_employee_pdf(request, employee_id):
    """
    Generate PDF Report for a single employee (Corporate Modern Style).
    """
    from attendance.models import Employee, AttendanceLog, EmployeeLeave
    from django.utils import timezone
    from datetime import date, timedelta, time, datetime
    import calendar
    from reportlab.lib.pagesizes import portrait, A4, LEGAL
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from attendance.utils import get_employee_schedule
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    try:
        month = int(request.GET.get('month'))
        year = int(request.GET.get('year'))
    except (TypeError, ValueError):
        now = timezone.now()
        month = now.month
        year = now.year
        
    start_date = date(year, month, 1)
    _, days_in_month = calendar.monthrange(year, month)
    end_date = date(year, month, days_in_month)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(LEGAL),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    # --- CORPORATE COLORS ---
    # Navy Blue: #1f3a93 (31, 58, 147)
    corpo_blue = colors.Color(31/255, 58/255, 147/255)
    corpo_grey = colors.Color(240/255, 240/255, 240/255) # Light Grey for rows
    corpo_red = colors.Color(231/255, 76/255, 60/255)
    corpo_green = colors.Color(46/255, 204/255, 113/255)
    corpo_orange = colors.Color(243/255, 156/255, 18/255)
    
    # --- HEADER FUNCTION (ON FIRST PAGE) ---
    def draw_header(canvas, doc):
        canvas.saveState()
        
        # REMOVED COMPANY HEADER AS REQUESTED
        
        # Footer Timestamp (All Pages)
        canvas.setFillColor(colors.grey)
        canvas.setFont("Helvetica", 8)
        ts = timezone.now().strftime("%d-%m-%Y %H:%M")
        canvas.drawRightString(19.5*cm, 1*cm, f"Dicetak pada: {ts}")
        
        canvas.restoreState()

    # --- FETCH DATA ---
    logs = AttendanceLog.objects.filter(
        employee=employee,
        timestamp__date__range=[start_date, end_date]
    ).order_by('timestamp')
    
    leaves = EmployeeLeave.objects.filter(
        employee=employee,
        start_date__lte=end_date,
        end_date__gte=start_date
    )
    leave_map = {}
    for l in leaves:
        curr = max(l.start_date, start_date)
        end = min(l.end_date, end_date)
        while curr <= end:
            leave_map[curr] = l.leave_type
            curr += timedelta(days=1)
            
    # Stats Init
    stats = {'H': 0, 'I': 0, 'S': 0, 'A': 0, 'L': 0}
    
    # --- PREPARE TABLE DATA (With Calculation Logic) ---
    # Determine Columns
    is_wa = (employee.attendance_method == 'WHATSAPP')
    
    if is_wa:
        headers = ['Tgl', 'Hari', 'Pagi', 'CP1', 'Isoma', 'CP2', 'Pulang', 'Status']
        col_widths = [1.8*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 3*cm] 
    else:
        headers = ['Tanggal', 'Hari', 'Masuk', 'Pulang', 'Telat', 'Lembur', 'Status']
        col_widths = [2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 4*cm]

    rows_buffer = []
    wa_missing = {'Pagi': 0, 'CP1': 0, 'Istirahat': 0, 'CP2': 0, 'Pulang': 0}
    
    import locale
    try:
        locale.setlocale(locale.LC_TIME, 'id_ID')
    except:
        pass

    import holidays
    id_holidays = holidays.Indonesia(years=year)

    for d in range(1, days_in_month + 1):
        day_date = date(year, month, d)
        day_str = day_date.strftime("%d")
        day_name = day_date.strftime("%a") # Short day name
        
        day_logs = logs.filter(timestamp__date=day_date)
        
        status = "-"
        status_color = colors.black
        
        leave_type = leave_map.get(day_date)
        row_vals = []
        
        if leave_type:
            status = leave_type
            stats[leave_type[0]] += 1
            if leave_type.startswith('C'): stats['I'] += 1 # Cuti -> Izin category
            row_vals = ["-"] * (len(headers) - 3)
            status_color = corpo_blue # Blue for Leave
            
        elif day_logs.exists():
            status = "HADIR"
            stats['H'] += 1
            
            if is_wa:
                # WA LOGIC
                l_pagi = day_logs.filter(log_category='MASUK').first()
                l_cp1  = day_logs.filter(log_category='CHECK1').first()
                l_isti = day_logs.filter(log_category='ISTIRAHAT').first()
                l_cp2  = day_logs.filter(log_category='CHECK2').first()
                l_plg  = day_logs.filter(log_category='PULANG').last()
                
                times = []
                # Pagi Check
                if l_pagi: 
                    t_str = timezone.localtime(l_pagi.timestamp).strftime("%H:%M")
                    times.append(t_str)
                    # Late Check (Hardcoded > 08:05 for WA as requested/legacy)
                    if timezone.localtime(l_pagi.timestamp).time() > time(8, 5):
                         status += " (TRL)"
                         stats['L'] += 1
                         status_color = corpo_orange
                else: 
                    times.append("-")
                    wa_missing['Pagi'] += 1
                
                # Other Checkpoints
                for l_chk, k in [(l_cp1, 'CP1'), (l_isti, 'Istirahat'), (l_cp2, 'CP2'), (l_plg, 'Pulang')]:
                    if l_chk: times.append(timezone.localtime(l_chk.timestamp).strftime("%H:%M"))
                    else: 
                        times.append("-")
                        wa_missing[k] += 1
                row_vals = times

            else:
                # FINGERPRINT LOGIC (With Dynamic Schedule Fix)
                l_in = day_logs.filter(log_category='MASUK').first()
                l_out = day_logs.filter(log_category='PULANG').last()
                
                t_in = timezone.localtime(l_in.timestamp).strftime("%H:%M") if l_in else "-"
                t_out = timezone.localtime(l_out.timestamp).strftime("%H:%M") if l_out else "-"
                
                # Telat Calculation
                telat_str = "-"
                
                # --- FIXED DYNAMIC LOGIC START ---
                daily_sch, _ = get_employee_schedule(employee, day_date)
                sch_in = time(8, 0)
                tol = 0
                if daily_sch:
                    sch_in = daily_sch.clock_in
                    tol = daily_sch.late_tolerance
                
                dt_sch = datetime.combine(date(2000,1,1), sch_in)
                dt_threshold = dt_sch + timedelta(minutes=tol)
                
                if l_in:
                    tm = timezone.localtime(l_in.timestamp).time()
                    if tm > dt_threshold.time():
                        dt_in = datetime.combine(date(2000,1,1), tm)
                        diff = (dt_in - dt_threshold).total_seconds() / 60
                        mins_late = int(diff)
                        if mins_late > 0:
                            telat_str = f"{mins_late}m"
                            stats['L'] += 1
                            status += " (TRL)"
                            status_color = corpo_orange
                # --- FIXED DYNAMIC LOGIC END ---
                
                # Lembur Logic (simplified)
                lembur_str = "-"
                work_end_hour = 12 if day_date.weekday() == 5 else 17
                if l_out and timezone.localtime(l_out.timestamp).time() > time(work_end_hour, 0):
                     tm = timezone.localtime(l_out.timestamp).time()
                     mins_over = (tm.hour * 60 + tm.minute) - (work_end_hour * 60)
                     if mins_over > 0:
                         lembur_str = f"+{mins_over}m"
                
                row_vals = [t_in, t_out, telat_str, lembur_str]
            
        else:
            if day_date.weekday() == 6:
                status = "LIBUR"
                status_color = colors.grey
            elif day_date in id_holidays:
                status = "LIBUR"
                status_color = corpo_red
            else:
                status = "ALPHA"
                stats['A'] += 1
                status_color = corpo_red
            
            row_vals = ["-"] * (len(headers) - 3)
            
        # Build Table Row
        # Tgl, Hari, ...data..., Status
        row = [f"{day_date.strftime('%d')}-{day_date.strftime('%m')}", day_name] + row_vals + [status]
        
        # Add Style for this specific row later? No, usually by index.
        # We'll store row data and styles separately or use conditional formatting in TableStyle
        rows_buffer.append((row, status_color)) # Tuple (data, color)

    # --- BUILD FLOWABLES ---
    elements = []
    
    # 1. Spacer for Header (REDUCED from 2.5cm to 1cm)
    elements.append(Spacer(1, 1*cm))
    
    # 2. Report Title
    styles = getSampleStyleSheet()
    # CENTERED TITLE
    title_style = ParagraphStyle(name='Title', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, textColor=corpo_blue, spaceAfter=2, alignment=TA_CENTER)
    elements.append(Paragraph("LAPORAN ABSENSI PERSONAL", title_style))
    
    # Check current locale for Month Name
    elements.append(Paragraph(f"Periode: {calendar.month_name[month]} {year}", ParagraphStyle(name='SubTitle', parent=styles['Normal'], alignment=TA_CENTER)))
    
    # REDUCED SPACER (Previously 0.8cm)
    elements.append(Spacer(1, 0.4*cm))
    
    # 3. Bio & Summary Cards (Side by Side)
    # Left: Bio Text
    # Right: Grid of 4 cards
    
    # Bio Table
    bio_data = [
        [Paragraph("<b>Nama</b>", styles['Normal']), Paragraph(f": {employee.full_name}", styles['Normal'])],
        [Paragraph("<b>NIK</b>", styles['Normal']), Paragraph(f": {employee.nik}", styles['Normal'])],
        [Paragraph("<b>Jabatan</b>", styles['Normal']), Paragraph(f": {employee.get_employee_type_display()}", styles['Normal'])],
        [Paragraph("<b>Lokasi</b>", styles['Normal']), Paragraph(f": {employee.home_base.name if employee.home_base else '-'}", styles['Normal'])],
    ]
    t_bio = Table(bio_data, colWidths=[2.5*cm, 5*cm])
    t_bio.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
    ]))
    
    # Cards Table
    # Helper to make a colored card
    def make_card(number, label, bg_color):
        s_card = ParagraphStyle(name='CardVal', parent=styles['Normal'], fontSize=16, textColor=colors.white, alignment=TA_CENTER, fontName='Helvetica-Bold')
        s_lbl = ParagraphStyle(name='CardLbl', parent=styles['Normal'], fontSize=8, textColor=colors.white, alignment=TA_CENTER)
        return Table([
            [Paragraph(str(number), s_card)],
            [Paragraph(label, s_lbl)]
        ], colWidths=[2.2*cm], rowHeights=[1*cm, 0.5*cm], style=TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), bg_color),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [5,5,5,5]), # Not supported in all reportlab versions, but ignored if not
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

    card_hadir = make_card(stats['H'], "HADIR", corpo_green)
    card_telat = make_card(stats['L'], "TELAT", corpo_orange)
    card_izin  = make_card(stats['S'] + stats['I'], "IZIN/SKT", corpo_blue)
    card_alpha = make_card(stats['A'], "ALPHA", corpo_red)
    
    t_cards = Table([
        [card_hadir, card_telat, card_izin, card_alpha]
    ], colWidths=[2.4*cm]*4)
    t_cards.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
    ]))
    
    # Master Layout Table (Bio | Cards)
    t_master = Table([
        [t_bio, t_cards]
    ], colWidths=[8*cm, 10*cm])
    t_master.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    
    elements.append(t_master)
    
    # REDUCED SPACER (Previously 0.8cm)
    elements.append(Spacer(1, 0.3*cm))
    
    # 4. Attendance Table
    final_rows = [headers] + [x[0] for x in rows_buffer]
    
    t = Table(final_rows, colWidths=col_widths)
    
    # Base Style
    tbl_style = [
        ('BACKGROUND', (0,0), (-1,0), corpo_blue), # Header BG
        ('TEXTCOLOR', (0,0), (-1,0), colors.white), # Header Text
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,0), 0, corpo_blue), # Header Line
        ('LINEBELOW', (0,1), (-1,-1), 0.5, colors.lightgrey), # Horizontal Lines
    ]
    
    # Zebra & Status Color
    for i, (row_data, text_color) in enumerate(rows_buffer):
        row_idx = i + 1 # Skip header
        if i % 2 == 1:
            tbl_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), corpo_grey))
        
        # Apply Text Color to Status Column (Last Col)
        tbl_style.append(('TEXTCOLOR', (-1, row_idx), (-1, row_idx), text_color))
        # Bold if Alpha or Late
        if text_color in [corpo_red, corpo_orange]:
             tbl_style.append(('FONTNAME', (-1, row_idx), (-1, row_idx), 'Helvetica-Bold'))

    t.setStyle(TableStyle(tbl_style))
    elements.append(t)
    
    # 5. Signatures (Wrapped in KeepTogether)
    sig_elements = []
    sig_elements.append(Spacer(1, 0.5*cm)) # Reduced from 1cm
    
    sig_style = ParagraphStyle(name='Sig', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9)
    # Layout using table
    ts = timezone.now().strftime("%d %B %Y")
    # Determine Month Name in ID if possible, otherwise English
    try:
        import locale
        locale.setlocale(locale.LC_TIME, 'id_ID')
        ts = timezone.now().strftime("%d %B %Y")
    except:
        pass
        
    sig_table = Table([
        [Paragraph(f"Jambi, {ts}", sig_style), ""],
        [Paragraph("Dibuat Oleh,", sig_style), Paragraph("Mengetahui,", sig_style)],
        ["", ""],
        ["", ""],
        ["", ""],
        [Paragraph("( Admin )", sig_style), Paragraph("( HRD Manager )", sig_style)]
    ], colWidths=[9*cm, 9*cm])
    
    sig_elements.append(sig_table)
    elements.append(KeepTogether(sig_elements))
    
    # Build
    doc.build(elements, onFirstPage=draw_header)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Laporan_{employee.full_name}_{calendar.month_name[month]}_{year}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response

@login_required
def admin_notification_list(request):
    """
    List all notifications for Admin monitoring.
    """
    # Authorization Check
    user_role = 'USER'
    if hasattr(request.user, 'employee_profile'):
        user_role = request.user.employee_profile.role
    
    if not (request.user.is_superuser or user_role == 'ADMIN'):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Hanya Admin yang dapat mengakses halaman ini.")
    
    from .models import Notification
    from django.core.paginator import Paginator
    
    # Query all notifications
    notif_list = Notification.objects.all().select_related('recipient').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(notif_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Notifikasi Sistem',
        'notifications': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'portal/admin_notification_list.html', context)
